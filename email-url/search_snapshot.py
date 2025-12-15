#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
search_snapshot.py

按 Excel 行构造搜索关键字，调用 ScrapingBee Google Search API，
取前 3 条 URL 做截图快照，并将过程记录到 log.csv 中。

依赖：pip install openpyxl requests scrapingbee

使用示例：
    python search_snapshot.py \
        --input-file=/path/to/file.xlsx \
        --sheet=Sheet1 \
        --search-columns=C*,D \
        --rows=3+ \
        --debug

配置文件：config.toml 或环境变量 SCRAPINGBEE_API_KEY
"""

# 导入标准库和第三方库
import argparse, csv, datetime as dt, hashlib, json, os, sys, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Any
import requests, urllib3
from openpyxl import load_workbook
from scrapingbee import ScrapingBeeClient

# 禁用 SSL 证书验证警告，避免在 HTTPS 请求时显示警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 兼容不同 Python 版本的 toml 库导入
try: import tomllib
except ImportError:
    try: import tomli as tomllib
    except ImportError: tomllib = None

# 全局常量定义
DEBUG = False  # 调试模式标志
DIRECT_DOWNLOAD_EXTENSIONS = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.zip', '.png', '.jpg', '.jpeg', '.gif'}  # 支持直接下载的文件扩展名
SEARCH_LOG_HEADER = ["sheet", "row", "keywords", "search_time", "search_duration_ms", "search_result_json", "search_error", "url1", "url2", "url3"]  # 搜索日志表头
SNAPSHOT_LOG_HEADER = ["url", "sheets", "rows", "keywords", "snapshot_path", "snapshot_error", "snapshot_time", "is_direct_download", "file_size_bytes"]  # 快照日志表头

# ==================== 工具函数 ====================

def log_print(*args, level="INFO", task_prefix="", **kwargs):
    """通用日志打印函数，带时间戳和级别标识"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] [{level}] {task_prefix}", *args, **kwargs)


# 配置和环境相关函数
def load_env_file(path: str = ".env"):
    """加载 .env 文件到环境变量"""
    if not os.path.exists(path): return
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line: continue
                key, value = line.split("=", 1)
                key, value = key.strip(), value.strip()
                if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
                    value = value[1:-1]
                if key and key not in os.environ: os.environ[key] = value
    except Exception as e: log_print("加载 .env 文件出错：", e)

def load_config(config_path: str) -> Dict[str, Any]:
    """读取配置，默认值兜底"""
    cfg = {"timeout_seconds": 120, "concurrency": 1, "retry_times": 1, "proxy": None}
    if not os.path.exists(config_path) or tomllib is None: return cfg
    try:
        with open(config_path, "rb") as f:
            data = tomllib.load(f)
            bee_cfg = data.get("scrapingbee", {})
            cfg.update({k: v for k, v in bee_cfg.items() if k in cfg})
    except Exception as e: log_print("解析 config.toml 出错，使用默认配置。错误：", e)
    return cfg

def column_letters_to_index(letters: str) -> int:
    """Excel 列字母转索引"""
    letters = letters.upper()
    result = 0
    for ch in letters:
        if not ("A" <= ch <= "Z"): raise ValueError(f"非法列字母: {letters}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result

def parse_search_columns(spec: str) -> List[Tuple[int, bool]]:
    """解析搜索列参数"""
    result = []
    for token in spec.split(","):
        token = token.strip()
        if not token: continue
        exact = token.endswith("*")
        col_letters = token[:-1] if exact else token
        col_index = column_letters_to_index(col_letters)
        result.append((col_index, exact))
    if not result: raise ValueError("search-columns 解析结果为空，请检查参数。")
    return result

def parse_rows_spec(spec: str, max_row: int) -> Tuple[int, int]:
    """解析行范围参数"""
    spec = spec.strip()
    if spec.endswith("+"):
        start_row = int(spec[:-1])
        end_row = max_row
    else:
        parts = spec.split("-", 1)
        start_row, end_row = int(parts[0]), int(parts[1])
    if start_row < 1 or end_row < start_row or start_row > max_row:
        raise ValueError(f"rows 范围非法：{start_row}-{end_row}")
    return start_row, min(end_row, max_row)

def sha1_hex(text: str) -> str: return hashlib.sha1(text.encode("utf-8")).hexdigest()

# ==================== ScrapingBee API 调用 ====================

def retry_operation(operation_func, *args, retry_times=1, task_prefix="", operation_name="", **kwargs):
    """通用重试函数"""
    errors = []
    attempts = 1 + max(retry_times, 0)
    op_name = operation_name or operation_func.__name__  # 使用显式名称或函数名，避免泄露敏感参数
    for attempt in range(1, attempts + 1):
        try:
            result = operation_func(*args, **kwargs)
            if isinstance(result, tuple) and len(result) >= 2 and result[1] is None:
                log_print(f"[成功] {op_name} | 尝试 {attempt}/{attempts}", task_prefix=task_prefix)
                return result
            elif isinstance(result, tuple) and len(result) >= 2:
                error = result[1]
            else:
                error = result
            log_print(f"[失败] {op_name} | 尝试 {attempt}/{attempts} | 错误: {error}", task_prefix=task_prefix)
            errors.append(f"attempt {attempt}: {error}")
        except Exception as e:
            log_print(f"[异常] {op_name} | 尝试 {attempt}/{attempts} | 错误: {e}", task_prefix=task_prefix)
            errors.append(f"attempt {attempt}: {e}")
        time.sleep(min(1.0 * attempt, 5.0))
    return None, errors

def search_google(api_key: str, keywords: str, timeout: int, proxies=None):
    """执行 Google 搜索"""
    url = "https://app.scrapingbee.com/api/v1/google"
    params = {"api_key": api_key, "search": keywords, "language": "en"}
    start = time.monotonic()
    try:
        response = requests.get(url, params=params, timeout=timeout, proxies=proxies)
        duration = time.monotonic() - start
        if response.status_code != 200:
            return None, f"HTTP {response.status_code}: {response.text[:1000]}", duration
        data = response.json()
        organic = data.get("organic_results", [])
        if isinstance(organic, list):
            return organic, None, duration
        return None, f"organic_results 非列表. 响应内容: {json.dumps(data, ensure_ascii=False)[:2000]}", duration
    except Exception as e:
        return None, f"请求异常: {e}", time.monotonic() - start

def take_screenshot(client: ScrapingBeeClient, url: str, save_path: str, timeout: int):
    """执行网页截图"""
    start = time.monotonic()
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        response = client.get(url, params={"screenshot": True, "screenshot_full_page": True, "transparent_status_code": True}, timeout=timeout)
        duration = time.monotonic() - start
        content = getattr(response, "content", None)
        if not content: return None, f"empty content for url={url}", duration
        with open(save_path, "wb") as f: f.write(content)
        file_size = len(content)
        log_print(f"[截图] ✓ 成功 | URL: {url} | 大小: {file_size/1024:.1f}KB | 耗时: {duration:.2f}s")
        return file_size, None, duration
    except Exception as e:
        return None, f"exception for url={url}: {e}", time.monotonic() - start

def download_direct(url: str, save_path: str, timeout: int, proxies=None):
    """执行直接下载"""
    start = time.monotonic()
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', 'Accept': '*/*'}
        response = requests.get(url, headers=headers, timeout=timeout, stream=True, verify=False, proxies=proxies, allow_redirects=True)
        duration = time.monotonic() - start
        if response.status_code != 200: return None, f"HTTP {response.status_code}", duration
        file_size = 0
        with open(save_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk: f.write(chunk); file_size += len(chunk)
        log_print(f"[下载] ✓ 成功 | URL: {url} | 大小: {file_size/1024:.1f}KB | 耗时: {duration:.2f}s")
        return file_size, None, duration
    except Exception as e:
        return None, f"下载异常: {e}", time.monotonic() - start

def is_direct_downloadable(url: str) -> bool:
    """检查是否可直接下载，根据文件扩展名判断"""
    try:
        from urllib.parse import urlparse, unquote
        path = unquote(urlparse(url).path).lower()
        return any(path.endswith(ext) for ext in DIRECT_DOWNLOAD_EXTENSIONS)
    except: return False

# ==================== 日志文件处理 ====================

def ensure_log_header(log_path: str, header: List[str]):
    """创建日志文件表头"""
    if not os.path.exists(log_path):
        with open(log_path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(header)

def load_existing_logs(search_log_path: str, snapshot_log_path: str, snapshot_root: str):
    """加载现有日志，返回缓存数据和完成行集合"""
    search_cache, snapshot_cache, row_done_set = {}, {}, set()

    # 读取搜索日志
    if os.path.exists(search_log_path):
        with open(search_log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                keywords = (row.get("keywords") or "").strip()
                if keywords and row.get("search_result_json"):
                    try: search_cache[keywords] = {"results": json.loads(row["search_result_json"]), "search_error": row.get("search_error", "")}
                    except: pass

    # 读取快照日志
    if os.path.exists(snapshot_log_path):
        with open(snapshot_log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = row.get("url", "")
                if url and row.get("snapshot_path") and not row.get("snapshot_error"):
                    full_path = os.path.join(snapshot_root, row["snapshot_path"])
                    if os.path.exists(full_path):
                        try: file_size = int(row.get("file_size_bytes", 0))
                        except: file_size = 0
                        snapshot_cache[url] = {
                            "snapshot_path": row["snapshot_path"],
                            "sheets": row.get("sheets", "").split("\n"),
                            "rows": row.get("rows", "").split("\n"),
                            "keywords": row.get("keywords", "").split("\n"),
                            "snapshot_error": "",
                            "snapshot_time": row.get("snapshot_time", ""),
                            "is_direct_download": row.get("is_direct_download", "").lower() == "true",
                            "file_size_bytes": file_size,
                        }

    # 读取搜索日志判断完成行
    if os.path.exists(search_log_path):
        with open(search_log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                sheet_name, row_no_str = row.get("sheet", ""), row.get("row", "")
                try: row_no = int(row_no_str)
                except: continue
                if row.get("search_error"): continue
                urls = [row.get(f"url{i}") for i in range(1, 4) if row.get(f"url{i}")]
                if not urls or all(snapshot_cache.get(url) for url in urls):
                    row_done_set.add((sheet_name, row_no))

    return search_cache, snapshot_cache, row_done_set

# ==================== 核心业务逻辑 ====================

def build_keywords(ws, row_idx: int, columns_spec: List[Tuple[int, bool]]) -> str:
    """构造搜索关键字"""
    parts = []
    for col_index, exact in columns_spec:
        cell = ws.cell(row=row_idx, column=col_index)
        if cell.value is None: continue
        s = str(cell.value).replace("\n", " ").strip()
        if not s: continue
        parts.append(f'"{s}"' if exact else s)
    return " ".join(parts).strip()

def process_row(sheet_name: str, row_idx: int, ws, columns_spec: List[Tuple[int, bool]], api_key: str, cfg: Dict[str, Any], snapshot_root: str, search_writer: csv.writer, snapshot_writer: csv.writer, search_cache: Dict, snapshot_cache: Dict, row_done_set: set, task_prefix: str = ""):
    """处理单行 Excel 数据：搜索关键词并截图快照"""
    # 检查是否已完成处理
    if (sheet_name, row_idx) in row_done_set: return True, False

    # 从 Excel 行构建搜索关键词
    keywords = build_keywords(ws, row_idx, columns_spec)
    if not keywords:
        # 空关键词直接标记完成
        search_writer.writerow([sheet_name, row_idx, keywords, dt.datetime.now().astimezone().isoformat(), 0, "", "empty keywords", "", "", ""])
        return True, False

    # 执行 Google 搜索逻辑
    timeout, retry_times = cfg["timeout_seconds"], cfg["retry_times"]
    proxies = {"http": cfg["proxy"], "https": cfg["proxy"]} if cfg.get("proxy") else None

    # 使用缓存或执行新搜索
    if keywords in search_cache:
        organic_results, search_error = search_cache[keywords]["results"], search_cache[keywords]["search_error"]
        search_duration_ms = 0
    else:
        result = retry_operation(search_google, api_key, keywords, timeout, retry_times=retry_times, proxies=proxies, task_prefix=task_prefix, operation_name=f"搜索: {keywords[:50]}")
        if result and result[0] is not None:
            organic_results, search_error, dur = result[0], "", result[2]
        else:
            errors = result[1] if result else ["unknown error"]
            organic_results, search_error, dur = [], "; ".join(errors) if isinstance(errors, list) else str(errors), 0
        search_duration_ms = int(dur * 1000)
        search_cache[keywords] = {"results": organic_results, "search_error": search_error}

    # 记录搜索结果到 CSV
    search_result_json = json.dumps(organic_results, ensure_ascii=False) if organic_results else ""
    urls = [item.get("url", "") for item in organic_results[:3] if isinstance(item, dict) and item.get("url")]
    urls.extend([""] * (3 - len(urls)))

    search_writer.writerow([sheet_name, row_idx, keywords, dt.datetime.now().astimezone().isoformat(), search_duration_ms, search_result_json, search_error, urls[0], urls[1], urls[2]])

    # 搜索失败则跳过快照处理
    if search_error: return False, True

    # 执行快照截图逻辑
    client = ScrapingBeeClient(api_key=api_key)
    has_errors = False

    # 处理每个搜索结果 URL 的快照
    for url in urls:
        if not url: continue

        # 检查是否已有成功的快照缓存
        if url in snapshot_cache and not snapshot_cache[url]["snapshot_error"]:
            # 更新现有缓存，追加当前行信息
            cache = snapshot_cache[url]
            cache["sheets"] = sorted(set(cache["sheets"] + [sheet_name]))
            cache["rows"] = sorted(set(cache["rows"] + [str(row_idx)]), key=lambda x: int(x) if x.isdigit() else 0)
            cache["keywords"] = sorted(set(cache["keywords"] + [keywords]))
            continue

        # 生成文件路径和确定下载方式
        h = sha1_hex(url)
        is_direct = is_direct_downloadable(url)

        # 根据下载方式确定文件扩展名
        if is_direct:
            try:
                from urllib.parse import urlparse, unquote
                path = unquote(urlparse(url).path)
                ext = os.path.splitext(path)[1].lower()
                if not ext or ext not in DIRECT_DOWNLOAD_EXTENSIONS: ext = ".pdf"
            except: ext = ".pdf"
        else: ext = ".png"

        # 构建文件路径：使用哈希值分层存储
        rel_path = os.path.join(h[:2], h[2:4], h[4:] + ext)
        full_path = os.path.join(snapshot_root, rel_path)
        snapshot_time = dt.datetime.now().astimezone().isoformat()

        # 根据下载方式选择处理函数
        if is_direct:
            # 直接下载文件
            result = retry_operation(download_direct, url, full_path, timeout, retry_times=retry_times, proxies=proxies, task_prefix=task_prefix, operation_name=f"下载: {url}")
            if result and result[0] is not None:
                file_size, _, _ = result
                snapshot_error = ""
            else:
                file_size, snapshot_error = 0, "; ".join(result[1]) if result else "download failed"
        else:
            # 使用 ScrapingBee 截图
            result = retry_operation(take_screenshot, client, url, full_path, timeout, retry_times=retry_times, task_prefix=task_prefix, operation_name=f"截图: {url}")
            if result and result[0] is not None:
                file_size, _, _ = result
                snapshot_error = ""
            else:
                file_size, snapshot_error = 0, "; ".join(result[1]) if result else "screenshot failed"

        # 记录是否有错误
        if snapshot_error: has_errors = True

        # 写入快照日志
        snapshot_writer.writerow([url, sheet_name, str(row_idx), keywords, rel_path if not snapshot_error else "", snapshot_error, snapshot_time if not snapshot_error else "", "true" if is_direct else "false", str(file_size) if not snapshot_error else ""])

        # 更新缓存（仅成功时）
        if not snapshot_error:
            snapshot_cache[url] = {
                "snapshot_path": rel_path, "sheets": [sheet_name], "rows": [str(row_idx)], "keywords": [keywords],
                "snapshot_error": "", "snapshot_time": snapshot_time, "is_direct_download": is_direct, "file_size_bytes": file_size
            }

    return not has_errors, has_errors

# ==================== 主程序入口 ====================

def main():
    """主程序入口，处理命令行参数并执行批量搜索快照任务"""
    global DEBUG

    # 参数解析
    parser = argparse.ArgumentParser(description="Excel 驱动的 ScrapingBee 搜索+快照脚本")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet", required=True, help="Sheet 名称")
    parser.add_argument("--search-columns", required=True, help="搜索列设置，例如 C*,D")
    parser.add_argument("--rows", required=True, help="行范围，例如 3+ 或 3-9")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    args = parser.parse_args()

    DEBUG = args.debug

    # 输入文件验证
    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path): log_print(f"输入文件不存在: {input_path}", level="ERROR"); sys.exit(1)

    # 路径和文件名处理
    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    snapshot_root = os.path.join(base_dir, f"{base_name}-snapshot")
    search_log_path = os.path.join(base_dir, f"{base_name}.search.csv")
    snapshot_log_path = os.path.join(base_dir, f"{base_name}.snapshot.csv")

    # 配置文件查找
    config_path = os.path.join(os.path.dirname(__file__), "config.toml")
    if not os.path.exists(config_path): config_path = os.path.join(base_dir, "config.toml")

    # 显示配置信息
    log_print("=" * 70, level="INFO")
    log_print("配置信息", level="INFO")
    log_print(f"  输入文件: {input_path}", level="INFO")
    log_print(f"  快照目录: {snapshot_root}", level="INFO")
    log_print(f"  搜索日志: {search_log_path}", level="INFO")
    log_print(f"  快照日志: {snapshot_log_path}", level="INFO")
    log_print(f"  配置文件: {config_path}", level="INFO")
    log_print("=" * 70, level="INFO")

    # 加载配置和环境变量
    load_env_file()
    cfg = load_config(config_path)

    # 代理设置
    if cfg.get("proxy"):
        os.environ["HTTP_PROXY"] = os.environ["HTTPS_PROXY"] = cfg["proxy"]
        log_print(f"代理设置: {cfg['proxy']}", level="INFO")

    # API 密钥验证
    api_key = os.environ.get("SCRAPINGBEE_API_KEY", "").strip()
    if not api_key: log_print("未找到 SCRAPINGBEE_API_KEY", level="ERROR"); sys.exit(1)

    # 解析搜索列参数
    columns_spec = parse_search_columns(args.search_columns)

    # 加载并验证 Excel 文件
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames: log_print(f"未找到 sheet: {args.sheet}", level="ERROR"); sys.exit(1)
    ws = wb[args.sheet]
    log_print(f"Sheet '{args.sheet}' 最大行号: {ws.max_row}", level="INFO")

    # 解析行范围参数
    start_row, end_row = parse_rows_spec(args.rows, ws.max_row)
    log_print(f"处理行范围: {start_row}-{end_row}", level="INFO")

    # 初始化日志文件
    ensure_log_header(search_log_path, SEARCH_LOG_HEADER)
    ensure_log_header(snapshot_log_path, SNAPSHOT_LOG_HEADER)

    # 加载现有日志和缓存
    search_cache, snapshot_cache, row_done_set = load_existing_logs(search_log_path, snapshot_log_path, snapshot_root)

    # 确定待处理任务
    tasks = [row for row in range(start_row, end_row + 1) if (args.sheet, row) not in row_done_set]
    if not tasks: log_print("指定范围内的行已全部处理完成", level="INFO"); return

    log_print(f"开始处理 | 总任务数: {len(tasks)}", level="INFO")

    # 设置并发数
    concurrency = max(1, cfg["concurrency"])
    log_print(f"并发线程数: {concurrency}", level="INFO")

    # 进度统计变量
    finished, success, failed = 0, 0, 0

    # 并发处理任务
    with open(search_log_path, "a", newline="", encoding="utf-8") as sf, \
         open(snapshot_log_path, "a", newline="", encoding="utf-8") as ssf:

        sw, ssw = csv.writer(sf), csv.writer(ssf)
        worker_id = 0

        def process_with_progress(row_idx):
            """处理单行任务并更新进度"""
            nonlocal worker_id, finished, success, failed
            worker_id += 1
            task_prefix = f"#{worker_id} - [{tasks.index(row_idx) + 1}/{len(tasks)}]"
            try:
                row_success, row_failed = process_row(args.sheet, row_idx, ws, columns_spec, api_key, cfg, snapshot_root, sw, ssw, search_cache, snapshot_cache, row_done_set, task_prefix)
                finished += 1
                if row_success: success += 1
                if row_failed: failed += 1
                percentage = (finished / len(tasks) * 100)
                log_print(f"{finished}/{len(tasks)} ({percentage:.1f}%) | 成功: {success} | 失败: {failed}", level="INFO", task_prefix=task_prefix)
            except Exception as e:
                log_print(f"行 {row_idx} 处理异常: {e}", level="ERROR", task_prefix=task_prefix)

        # 启动线程池执行任务
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            list(executor.map(process_with_progress, tasks))

    # 显示最终统计结果
    log_print("=" * 70, level="INFO")
    log_print("任务完成", level="INFO")
    log_print(f"  总任务: {len(tasks)}", level="INFO")
    log_print(f"  成功: {success}", level="INFO")
    log_print(f"  失败: {failed}", level="INFO")
    log_print("=" * 70, level="INFO")


if __name__ == "__main__":
    main()
