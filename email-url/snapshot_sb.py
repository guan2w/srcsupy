#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
snapshot_sb.py

从 Excel 文件读取 URL 列，调用 ScrapingBee 进行截图和 HTML 快照，
输出结果到新的 Excel 文件中。

依赖：pip install openpyxl requests scrapingbee

使用示例：
    python snapshot_sb.py \
        --excel=/path/to/file.xlsx \
        --sheet=Sheet1 \
        --url-columns=url1,url2 \
        --title-row=2 \
        --data-rows=3+ \
        --output=/path/to/file-snapshot.xlsx \
        --debug

配置文件：config.toml（同目录或脚本目录）+ 环境变量 SCRAPINGBEE_API_KEY
"""

# 导入标准库和第三方库
import argparse
import datetime as dt
import hashlib
import os
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Any, Optional

import requests
import urllib3
from openpyxl import Workbook, load_workbook
from scrapingbee import ScrapingBeeClient

# 禁用 SSL 证书验证警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 兼容不同 Python 版本的 toml 库导入
try:
    import tomllib
except ImportError:
    try:
        import tomli as tomllib
    except ImportError:
        tomllib = None

# 全局常量定义
DEBUG = False  # 调试模式标志

# 输出 Excel 表头
OUTPUT_HEADER = [
    "original_url",       # 原始 URL
    "source_sheet",       # 来源 sheet
    "source_row",         # 来源行号
    "source_column",      # 来源列名
    "snapshot_status",    # 快照状态: success / skipped / failed
    "snapshot_image",     # 截图文件相对路径
    "snapshot_html",      # HTML 文件相对路径
    "snapshot_time",      # 快照时间 (ISO8601)
    "duration_ms",        # 快照耗时 (毫秒)
    "image_size_bytes",   # 截图文件大小 (字节)
    "html_size_bytes",    # HTML 文件大小 (字节)
    "error_message",      # 错误信息
]

# ==================== 工具函数 ====================

def log_print(*args, level="INFO", task_prefix="", **kwargs):
    """通用日志打印函数，带时间戳和级别标识"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    prefix = f"[{now}] [{level}]"
    if task_prefix:
        prefix = f"{prefix} {task_prefix}"
    print(prefix, *args, **kwargs)


def debug_print(*args, **kwargs):
    """调试打印，仅在 DEBUG 模式下输出"""
    if DEBUG:
        log_print(*args, level="DEBUG", **kwargs)


# ==================== 配置和环境相关函数 ====================

def load_env_file(path: str = ".env"):
    """加载 .env 文件到环境变量"""
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key, value = key.strip(), value.strip()
                # 去除引号
                if (value.startswith('"') and value.endswith('"')) or \
                   (value.startswith("'") and value.endswith("'")):
                    value = value[1:-1]
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception as e:
        log_print(f"加载 .env 文件出错：{e}", level="WARN")


def load_config(config_path: str) -> Dict[str, Any]:
    """读取配置文件，返回合并后的配置字典"""
    # 默认配置
    cfg = {
        "timeout_seconds": 120,
        "concurrency": 200,
        "retry_times": 3,
        "proxy": None,
        "dom_html": True,
        "screenshot": True,
    }
    
    if not os.path.exists(config_path):
        log_print(f"配置文件不存在: {config_path}，使用默认配置", level="WARN")
        return cfg
    
    if tomllib is None:
        log_print("未安装 tomllib/tomli，无法解析配置文件，使用默认配置", level="WARN")
        return cfg
    
    try:
        with open(config_path, "rb") as f:
            data = tomllib.load(f)
            
        # 读取 [scrapingbee] 块
        bee_cfg = data.get("scrapingbee", {})
        for key in ["timeout_seconds", "concurrency", "retry_times", "proxy"]:
            if key in bee_cfg:
                cfg[key] = bee_cfg[key]
        
        # 读取 [snapshot] 块
        snapshot_cfg = data.get("snapshot", {})
        for key in ["dom_html", "screenshot"]:
            if key in snapshot_cfg:
                cfg[key] = snapshot_cfg[key]
                
        log_print(f"配置加载成功: {config_path}")
        debug_print(f"配置内容: {cfg}")
        
    except Exception as e:
        log_print(f"解析配置文件出错，使用默认配置。错误：{e}", level="WARN")
    
    return cfg


# ==================== 参数解析函数 ====================

def parse_data_rows(spec: str, max_row: int) -> Tuple[int, int]:
    """
    解析 data-rows 参数，支持格式：
    - "3+" 表示从第 3 行到最后
    - "3-5" 表示第 3 到第 5 行
    - "3" 表示仅第 3 行
    
    返回 (start_row, end_row)，均为 1-based
    """
    spec = spec.strip()
    
    if spec.endswith("+"):
        # 格式: 3+
        start_row = int(spec[:-1])
        end_row = max_row
    elif "-" in spec:
        # 格式: 3-5
        parts = spec.split("-", 1)
        start_row = int(parts[0])
        end_row = int(parts[1])
    else:
        # 格式: 3 (单行)
        start_row = end_row = int(spec)
    
    # 验证范围
    if start_row < 1:
        raise ValueError(f"起始行号必须 >= 1，当前: {start_row}")
    if end_row < start_row:
        raise ValueError(f"结束行号必须 >= 起始行号，当前: {start_row}-{end_row}")
    if start_row > max_row:
        raise ValueError(f"起始行号超出范围，最大行号: {max_row}，当前: {start_row}")
    
    return start_row, min(end_row, max_row)


def parse_url_columns(spec: str) -> List[str]:
    """
    解析 url-columns 参数，返回列名列表
    例如: "url1,url2" -> ["url1", "url2"]
    """
    columns = []
    for col in spec.split(","):
        col = col.strip()
        if col:
            columns.append(col)
    
    if not columns:
        raise ValueError("url-columns 参数不能为空")
    
    return columns


def sha1_hex(text: str) -> str:
    """计算 SHA1 哈希值"""
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


def get_snapshot_paths(url: str, snapshot_dir: str) -> Tuple[str, str, str, str]:
    """
    根据 URL 生成快照文件路径
    返回: (相对路径前缀, 完整PNG路径, 完整HTML路径, 相对PNG路径, 相对HTML路径)
    """
    h = sha1_hex(url)
    # 分层目录: hash[:2] / hash[2:4] / hash[4:]
    rel_dir = os.path.join(h[:2], h[2:4])
    base_name = h[4:]
    
    rel_png = os.path.join(rel_dir, f"{base_name}.png")
    rel_html = os.path.join(rel_dir, f"{base_name}.html")
    
    full_png = os.path.join(snapshot_dir, rel_png)
    full_html = os.path.join(snapshot_dir, rel_html)
    
    return full_png, full_html, rel_png, rel_html


# ==================== ScrapingBee API 调用 ====================

def take_snapshot(
    client: ScrapingBeeClient,
    url: str,
    save_png_path: str,
    save_html_path: str,
    timeout: int,
    do_screenshot: bool = True,
    do_html: bool = True,
    proxies: Optional[Dict] = None,
) -> Tuple[Optional[int], Optional[int], Optional[str], float]:
    """
    执行网页快照（截图和/或 HTML）
    
    返回: (png_size, html_size, error_message, duration_seconds)
    - 成功时 error_message 为 None
    - 失败时 png_size/html_size 可能为 None
    """
    start = time.monotonic()
    png_size = None
    html_size = None
    error = None
    
    # 创建目录
    os.makedirs(os.path.dirname(save_png_path), exist_ok=True)
    
    try:
        # 构建请求参数
        params = {
            "transparent_status_code": True,
            "wait": 3000,  # 等待页面加载
        }
        
        if do_screenshot:
            params["screenshot"] = True
            params["screenshot_full_page"] = True
        
        # 执行请求
        response = client.get(url, params=params, timeout=timeout)
        
        # 检查响应状态
        status_code = getattr(response, "status_code", None)
        if status_code and status_code >= 400:
            error = f"HTTP {status_code}"
            return png_size, html_size, error, time.monotonic() - start
        
        content = getattr(response, "content", None)
        if not content:
            error = "响应内容为空"
            return png_size, html_size, error, time.monotonic() - start
        
        # 处理截图
        if do_screenshot and content:
            # 先写入临时文件
            tmp_path = save_png_path + ".tmp"
            with open(tmp_path, "wb") as f:
                f.write(content)
            # 原子性重命名
            os.rename(tmp_path, save_png_path)
            png_size = len(content)
        
        # 处理 HTML（需要单独请求）
        if do_html:
            html_response = client.get(url, params={
                "transparent_status_code": True,
                "wait": 3000,
                "render_js": True,
            }, timeout=timeout)
            
            html_content = getattr(html_response, "content", None)
            if html_content:
                # 先写入临时文件
                tmp_path = save_html_path + ".tmp"
                with open(tmp_path, "wb") as f:
                    f.write(html_content)
                # 原子性重命名
                os.rename(tmp_path, save_html_path)
                html_size = len(html_content)
        
    except Exception as e:
        error = str(e)
    
    duration = time.monotonic() - start
    return png_size, html_size, error, duration


def retry_snapshot(
    client: ScrapingBeeClient,
    url: str,
    save_png_path: str,
    save_html_path: str,
    timeout: int,
    retry_times: int,
    do_screenshot: bool,
    do_html: bool,
    proxies: Optional[Dict],
    task_prefix: str = "",
) -> Tuple[Optional[int], Optional[int], Optional[str], float]:
    """带重试的快照函数"""
    errors = []
    total_duration = 0
    
    for attempt in range(1, retry_times + 1):
        png_size, html_size, error, duration = take_snapshot(
            client, url, save_png_path, save_html_path, timeout,
            do_screenshot, do_html, proxies
        )
        total_duration += duration
        
        if error is None:
            debug_print(f"快照成功 | URL: {url[:50]} | 尝试: {attempt}", task_prefix=task_prefix)
            return png_size, html_size, None, total_duration
        
        errors.append(f"尝试{attempt}: {error}")
        log_print(f"快照失败 | URL: {url[:50]} | 尝试: {attempt}/{retry_times} | 错误: {error}", 
                  level="WARN", task_prefix=task_prefix)
        
        if attempt < retry_times:
            time.sleep(min(1.0 * attempt, 5.0))
    
    return None, None, "; ".join(errors), total_duration


# ==================== Excel 读写函数 ====================

def read_urls_from_excel(
    excel_path: str,
    sheet_name: str,
    url_columns: List[str],
    title_row: int,
    start_row: int,
    end_row: int,
) -> List[Dict[str, Any]]:
    """
    从 Excel 读取 URL 列表
    返回: [{"url": ..., "sheet": ..., "row": ..., "column": ...}, ...]
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' 不存在。可用: {wb.sheetnames}")
    
    ws = wb[sheet_name]
    
    # 读取表头行，建立列名到索引的映射
    header_row = list(ws.iter_rows(min_row=title_row, max_row=title_row, values_only=True))[0]
    col_name_to_idx = {}
    for idx, name in enumerate(header_row, start=1):
        if name:
            col_name_to_idx[str(name).strip()] = idx
    
    debug_print(f"表头列名: {list(col_name_to_idx.keys())}")
    
    # 验证 URL 列存在
    missing_cols = [col for col in url_columns if col not in col_name_to_idx]
    if missing_cols:
        raise ValueError(f"URL 列不存在: {missing_cols}。可用列: {list(col_name_to_idx.keys())}")
    
    # 读取数据
    url_list = []
    seen_urls = set()  # 用于去重
    
    for row_idx in range(start_row, end_row + 1):
        row_data = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        for col_name in url_columns:
            col_idx = col_name_to_idx[col_name]
            cell_value = row_data[col_idx - 1] if col_idx <= len(row_data) else None
            
            if cell_value:
                url = str(cell_value).strip()
                if url and url.startswith(("http://", "https://")):
                    if url not in seen_urls:
                        seen_urls.add(url)
                        url_list.append({
                            "url": url,
                            "sheet": sheet_name,
                            "row": row_idx,
                            "column": col_name,
                        })
    
    wb.close()
    return url_list


def write_output_excel(output_path: str, results: List[Dict[str, Any]]):
    """写入输出 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshot Results"
    
    # 写入表头
    for col_idx, header in enumerate(OUTPUT_HEADER, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 写入数据
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.get("url", ""))
        ws.cell(row=row_idx, column=2, value=result.get("sheet", ""))
        ws.cell(row=row_idx, column=3, value=result.get("row", ""))
        ws.cell(row=row_idx, column=4, value=result.get("column", ""))
        ws.cell(row=row_idx, column=5, value=result.get("status", ""))
        ws.cell(row=row_idx, column=6, value=result.get("image_path", ""))
        ws.cell(row=row_idx, column=7, value=result.get("html_path", ""))
        ws.cell(row=row_idx, column=8, value=result.get("snapshot_time", ""))
        ws.cell(row=row_idx, column=9, value=result.get("duration_ms", 0))
        ws.cell(row=row_idx, column=10, value=result.get("image_size", 0))
        ws.cell(row=row_idx, column=11, value=result.get("html_size", 0))
        ws.cell(row=row_idx, column=12, value=result.get("error", ""))
    
    wb.save(output_path)
    log_print(f"输出文件已保存: {output_path}")


# ==================== 主程序 ====================

def process_url(
    task: Dict[str, Any],
    client: ScrapingBeeClient,
    snapshot_dir: str,
    cfg: Dict[str, Any],
    task_idx: int,
    total_tasks: int,
) -> Dict[str, Any]:
    """处理单个 URL 的快照任务"""
    url = task["url"]
    task_prefix = f"[{task_idx}/{total_tasks}]"
    
    # 获取快照文件路径
    full_png, full_html, rel_png, rel_html = get_snapshot_paths(url, snapshot_dir)
    
    # 检查是否已存在（断点续传）
    do_screenshot = cfg["screenshot"]
    do_html = cfg["dom_html"]
    
    png_exists = os.path.exists(full_png) if do_screenshot else True
    html_exists = os.path.exists(full_html) if do_html else True
    
    if png_exists and html_exists:
        # 已存在，跳过
        png_size = os.path.getsize(full_png) if do_screenshot and os.path.exists(full_png) else 0
        html_size = os.path.getsize(full_html) if do_html and os.path.exists(full_html) else 0
        
        debug_print(f"跳过已存在 | URL: {url[:50]}", task_prefix=task_prefix)
        
        return {
            "url": url,
            "sheet": task["sheet"],
            "row": task["row"],
            "column": task["column"],
            "status": "skipped",
            "image_path": rel_png if do_screenshot else "",
            "html_path": rel_html if do_html else "",
            "snapshot_time": "",
            "duration_ms": 0,
            "image_size": png_size,
            "html_size": html_size,
            "error": "",
        }
    
    # 执行快照
    log_print(f"开始快照 | URL: {url[:60]}", task_prefix=task_prefix)
    
    proxies = None
    if cfg.get("proxy"):
        proxies = {"http": cfg["proxy"], "https": cfg["proxy"]}
    
    png_size, html_size, error, duration = retry_snapshot(
        client=client,
        url=url,
        save_png_path=full_png,
        save_html_path=full_html,
        timeout=cfg["timeout_seconds"],
        retry_times=cfg["retry_times"],
        do_screenshot=do_screenshot,
        do_html=do_html,
        proxies=proxies,
        task_prefix=task_prefix,
    )
    
    snapshot_time = dt.datetime.now().astimezone().isoformat()
    
    if error:
        log_print(f"快照失败 | URL: {url[:50]} | 错误: {error}", level="ERROR", task_prefix=task_prefix)
        status = "failed"
    else:
        log_print(f"快照成功 | URL: {url[:50]} | PNG: {png_size}B, HTML: {html_size}B | 耗时: {duration:.2f}s", 
                  task_prefix=task_prefix)
        status = "success"
    
    return {
        "url": url,
        "sheet": task["sheet"],
        "row": task["row"],
        "column": task["column"],
        "status": status,
        "image_path": rel_png if do_screenshot and png_size else "",
        "html_path": rel_html if do_html and html_size else "",
        "snapshot_time": snapshot_time,
        "duration_ms": int(duration * 1000),
        "image_size": png_size or 0,
        "html_size": html_size or 0,
        "error": error or "",
    }


def main():
    """主程序入口"""
    global DEBUG
    
    # 参数解析
    parser = argparse.ArgumentParser(
        description="Excel URL 快照工具 - 使用 ScrapingBee 进行网页截图和 HTML 快照"
    )
    parser.add_argument("--excel", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet", required=True, help="Sheet 名称")
    parser.add_argument("--url-columns", required=True, help="URL 列名，逗号分隔，如: url1,url2")
    parser.add_argument("--title-row", type=int, default=1, help="表头行号 (1-based，默认 1)")
    parser.add_argument("--data-rows", required=True, help="数据行范围，如: 3+ 或 3-5")
    parser.add_argument("--output", help="输出 Excel 文件路径 (默认: {input}-snapshot.xlsx)")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    args = parser.parse_args()
    
    DEBUG = args.debug
    
    # 验证输入文件
    excel_path = os.path.abspath(args.excel)
    if not os.path.exists(excel_path):
        log_print(f"输入文件不存在: {excel_path}", level="ERROR")
        sys.exit(1)
    
    # 路径处理
    base_dir = os.path.dirname(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    snapshot_dir = os.path.join(base_dir, f"{base_name}-snapshot")
    
    # 输出文件路径
    if args.output:
        output_path = os.path.abspath(args.output)
    else:
        output_path = os.path.join(base_dir, f"{base_name}-snapshot.xlsx")
    
    # 查找配置文件
    config_path = os.path.join(os.path.dirname(__file__), "config.toml")
    if not os.path.exists(config_path):
        config_path = os.path.join(base_dir, "config.toml")
    
    # 显示配置信息
    log_print("=" * 70)
    log_print("配置信息")
    log_print(f"  输入文件: {excel_path}")
    log_print(f"  Sheet: {args.sheet}")
    log_print(f"  URL 列: {args.url_columns}")
    log_print(f"  表头行: {args.title_row}")
    log_print(f"  数据行: {args.data_rows}")
    log_print(f"  快照目录: {snapshot_dir}")
    log_print(f"  输出文件: {output_path}")
    log_print(f"  配置文件: {config_path}")
    log_print("=" * 70)
    
    # 加载环境变量和配置
    load_env_file()
    load_env_file(os.path.join(base_dir, ".env"))
    cfg = load_config(config_path)
    
    # 代理设置
    if cfg.get("proxy"):
        os.environ["HTTP_PROXY"] = cfg["proxy"]
        os.environ["HTTPS_PROXY"] = cfg["proxy"]
        log_print(f"代理设置: {cfg['proxy']}")
    
    # API 密钥验证
    api_key = os.environ.get("SCRAPINGBEE_API_KEY", "").strip()
    if not api_key:
        log_print("未找到 SCRAPINGBEE_API_KEY，请设置环境变量或 .env 文件", level="ERROR")
        sys.exit(1)
    
    # 解析参数
    url_columns = parse_url_columns(args.url_columns)
    log_print(f"URL 列: {url_columns}")
    
    # 加载 Excel，获取最大行号
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        log_print(f"Sheet '{args.sheet}' 不存在。可用: {wb.sheetnames}", level="ERROR")
        sys.exit(1)
    ws = wb[args.sheet]
    max_row = ws.max_row
    wb.close()
    
    log_print(f"Sheet '{args.sheet}' 最大行号: {max_row}")
    
    # 解析数据行范围
    start_row, end_row = parse_data_rows(args.data_rows, max_row)
    log_print(f"处理数据行范围: {start_row} - {end_row}")
    
    # 读取 URL 列表
    url_list = read_urls_from_excel(
        excel_path, args.sheet, url_columns, 
        args.title_row, start_row, end_row
    )
    
    if not url_list:
        log_print("未找到有效的 URL", level="WARN")
        sys.exit(0)
    
    log_print(f"共读取 {len(url_list)} 个唯一 URL（已去重）")
    
    # 创建快照目录
    os.makedirs(snapshot_dir, exist_ok=True)
    
    # 初始化 ScrapingBee 客户端
    client = ScrapingBeeClient(api_key=api_key)
    
    # 并发配置
    concurrency = min(cfg["concurrency"], 500)
    log_print(f"并发线程数: {concurrency}")
    log_print(f"截图: {'启用' if cfg['screenshot'] else '禁用'}")
    log_print(f"HTML: {'启用' if cfg['dom_html'] else '禁用'}")
    log_print("=" * 70)
    
    # 统计变量
    results = []
    lock = threading.Lock()
    finished = 0
    success = 0
    skipped = 0
    failed = 0
    
    def process_with_progress(task_and_idx):
        """处理任务并更新进度"""
        nonlocal finished, success, skipped, failed
        task, idx = task_and_idx
        
        result = process_url(task, client, snapshot_dir, cfg, idx, len(url_list))
        
        with lock:
            results.append(result)
            finished += 1
            if result["status"] == "success":
                success += 1
            elif result["status"] == "skipped":
                skipped += 1
            else:
                failed += 1
            
            percentage = finished / len(url_list) * 100
            log_print(f"进度: {finished}/{len(url_list)} ({percentage:.1f}%) | "
                      f"成功: {success} | 跳过: {skipped} | 失败: {failed}")
    
    # 启动并发处理
    tasks_with_idx = [(task, idx + 1) for idx, task in enumerate(url_list)]
    
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        list(executor.map(process_with_progress, tasks_with_idx))
    
    # 按原始顺序排序结果
    results.sort(key=lambda x: (x["sheet"], x["row"], x["column"]))
    
    # 写入输出 Excel
    write_output_excel(output_path, results)
    
    # 显示最终统计
    log_print("=" * 70)
    log_print("任务完成")
    log_print(f"  总 URL 数: {len(url_list)}")
    log_print(f"  成功: {success}")
    log_print(f"  跳过: {skipped}")
    log_print(f"  失败: {failed}")
    log_print(f"  输出文件: {output_path}")
    log_print("=" * 70)


if __name__ == "__main__":
    main()
