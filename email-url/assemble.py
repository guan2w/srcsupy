#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
assemble.py

从输入文件、搜索日志、快照日志综合生成汇总 Excel 文件。

依赖：
    pip install openpyxl

使用示例：
    python assemble.py \
        --input-file=/path/to/file.xlsx \
        --sheet=Sheet1 \
        --search-columns=C*,D \
        --rows=3+ \
        --snapshot-prefix="http://192.168.51.109/snapshot/"

输出文件：
    {base_name}-数据行数-YYMMDD.hhmmss.xlsx
"""

import argparse
import csv
import datetime as dt
import json
import os
import shutil
import sys
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import load_workbook, Workbook


# ----------------- 全局状态 -----------------

DEBUG = False

def debug_print(*args, **kwargs):
    if DEBUG:
        now = dt.datetime.now().strftime("%H:%M:%S")
        print(f"[{now}] [DEBUG]", *args, **kwargs)


def column_letters_to_index(letters: str) -> int:
    """将 Excel 列字母（如 'A', 'C', 'AA'）转换为 1-based 列索引"""
    letters = letters.upper()
    result = 0
    for ch in letters:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"非法列字母: {letters}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def parse_search_columns(spec: str) -> List[Tuple[int, bool]]:
    """
    解析 --search-columns 参数，如 "C*,D,AA*"
    返回列表 [(col_index, exact_match), ...]
    exact_match=True 表示需要加双引号
    """
    result: List[Tuple[int, bool]] = []
    for token in spec.split(","):
        token = token.strip()
        if not token:
            continue
        exact = token.endswith("*")
        col_letters = token[:-1] if exact else token
        col_index = column_letters_to_index(col_letters)
        result.append((col_index, exact))
    if not result:
        raise ValueError("search-columns 解析结果为空，请检查参数。")
    return result


def parse_rows_spec(spec: str, max_row: int) -> Tuple[int, int]:
    """
    解析 --rows 参数：
        '3+'  -> (3, max_row)
        '3-9' -> (3, 9)
    返回 (start_row, end_row)，均为闭区间
    """
    spec = spec.strip()
    if spec.endswith("+"):
        start = spec[:-1]
        if not start.isdigit():
            raise ValueError(f"rows 参数非法：{spec}")
        start_row = int(start)
        end_row = max_row
    else:
        if "-" not in spec:
            raise ValueError(f"rows 参数非法：{spec}")
        parts = spec.split("-", 1)
        if not (parts[0].isdigit() and parts[1].isdigit()):
            raise ValueError(f"rows 参数非法：{spec}")
        start_row = int(parts[0])
        end_row = int(parts[1])
    if start_row < 1 or end_row < start_row:
        raise ValueError(f"rows 范围非法：{start_row}-{end_row}")
    # 同时不允许超过 max_row
    if start_row > max_row:
        raise ValueError(f"rows 起始行 {start_row} 大于 sheet 最大行 {max_row}")
    end_row = min(end_row, max_row)
    return start_row, end_row


def build_keywords_from_row(ws, row_idx: int, columns_spec: List[Tuple[int, bool]]) -> str:
    """
    从给定行构造搜索关键字字符串。
    columns_spec: [(col_index, exact), ...]
    """
    parts: List[str] = []
    for col_index, exact in columns_spec:
        cell = ws.cell(row=row_idx, column=col_index)
        value = cell.value
        if value is None:
            continue
        s = str(value)
        s = s.replace("\n", " ").strip()
        if not s:
            continue
        if exact:
            parts.append(f'"{s}"')
        else:
            parts.append(s)
    keywords = " ".join(parts).strip()
    return keywords


def load_search_log(search_log_path: str, sheet_name: str) -> Dict[str, Dict[str, Any]]:
    """
    读取搜索日志，返回以 keywords 为键的字典
    格式: {keywords: {row: int, search_time: str, search_duration_ms: int,
                      search_result_json: str, search_error: str,
                      urls: [url1, url2, url3]}}
    """
    search_data: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(search_log_path):
        print(f"警告：搜索日志文件不存在: {search_log_path}")
        return search_data

    with open(search_log_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("sheet") != sheet_name:
                continue

            keywords = (row.get("keywords") or "").strip()
            if not keywords:
                continue

            row_num = row.get("row", "")
            try:
                row_num = int(row_num)
            except ValueError:
                continue

            search_data[keywords] = {
                "row": row_num,
                "search_time": row.get("search_time", ""),
                "search_duration_ms": int(row.get("search_duration_ms", 0)),
                "search_result_json": row.get("search_result_json", ""),
                "search_error": row.get("search_error", ""),
                "urls": [
                    row.get("url1", ""),
                    row.get("url2", ""),
                    row.get("url3", ""),
                ]
            }

    print(f"从搜索日志加载了 {len(search_data)} 条记录")
    return search_data


def load_snapshot_log(snapshot_log_path: str) -> Dict[str, Dict[str, Any]]:
    """
    读取快照日志，返回以 url 为键的字典
    格式: {url: {snapshot_path: str, snapshot_error: str, ...}}
    """
    snapshot_data: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(snapshot_log_path):
        print(f"警告：快照日志文件不存在: {snapshot_log_path}")
        return snapshot_data

    with open(snapshot_log_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = row.get("url", "").strip()
            if not url:
                continue

            # 如果同一个URL有多条记录，保留最新的（最后出现的）
            snapshot_data[url] = {
                "snapshot_path": row.get("snapshot_path", ""),
                "snapshot_error": row.get("snapshot_error", ""),
                "snapshot_time": row.get("snapshot_time", ""),
                "is_direct_download": row.get("is_direct_download", "").lower() == "true",
                "file_size_bytes": int(row.get("file_size_bytes", 0)) if row.get("file_size_bytes") else 0,
            }

    print(f"从快照日志加载了 {len(snapshot_data)} 条记录")
    return snapshot_data


def calculate_snapshot_status(search_result_json: str, urls: List[str], snapshot_data: Dict[str, Dict[str, Any]]) -> Tuple[str, str]:
    """
    计算快照状态和错误信息
    返回 (status, errors)
    status 格式如 "2/3" 或 "0/3"
    errors 用换行符分隔的错误信息
    """
    # 解析搜索结果
    try:
        if search_result_json:
            results = json.loads(search_result_json)
            expected_count = min(len(results), 3)  # 最多3个
        else:
            expected_count = 0
    except json.JSONDecodeError:
        expected_count = 0

    if expected_count == 0:
        return "0/0", "搜索无结果"

    # 统计成功的快照数量
    successful_count = 0
    errors = []

    for url in urls[:expected_count]:
        if not url:
            continue

        snap_info = snapshot_data.get(url)
        if snap_info and snap_info.get("snapshot_path") and not snap_info.get("snapshot_error"):
            successful_count += 1
        elif snap_info and snap_info.get("snapshot_error"):
            errors.append(f"{url}: {snap_info['snapshot_error']}")
        else:
            errors.append(f"{url}: 尚未快照")

    status = f"{successful_count}/{expected_count}"
    error_text = "\n".join(errors) if errors else ""

    return status, error_text


def generate_output_filename(input_path: str, data_row_count: int) -> str:
    """生成输出文件名: {base_name}-数据行数-YYMMDD.hhmmss.xlsx"""
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    now = dt.datetime.now()
    timestamp = now.strftime("%y%m%d.%H%M%S")
    return f"{base_name}-{data_row_count}-{timestamp}.xlsx"


def main():
    parser = argparse.ArgumentParser(description="综合日志生成汇总 Excel 文件")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet", required=True, help="Sheet 名称")
    parser.add_argument("--search-columns", required=True, help="搜索列设置，例如 C*,D")
    parser.add_argument("--rows", required=True, help="行范围，例如 3+ 或 3-9")
    parser.add_argument("--snapshot-prefix", required=True, help="快照文件路径前缀")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    args = parser.parse_args()

    global DEBUG
    DEBUG = bool(args.debug)

    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path):
        print(f"错误：输入文件不存在: {input_path}")
        sys.exit(1)

    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]

    # 日志文件路径
    search_log_path = os.path.join(base_dir, f"{base_name}.search.csv")
    snapshot_log_path = os.path.join(base_dir, f"{base_name}.snapshot.csv")

    print("=" * 70)
    print("配置信息")
    print(f"  输入文件: {input_path}")
    print(f"  Sheet: {args.sheet}")
    print(f"  搜索列: {args.search_columns}")
    print(f"  行范围: {args.rows}")
    print(f"  快照前缀: {args.snapshot_prefix}")
    print(f"  搜索日志: {search_log_path}")
    print(f"  快照日志: {snapshot_log_path}")
    print("=" * 70)

    # 1. 解析参数
    try:
        columns_spec = parse_search_columns(args.search_columns)
        print(f"解析搜索列: {columns_spec}")
    except Exception as e:
        print(f"错误：解析 search-columns 出错: {e}")
        sys.exit(1)

    # 2. 打开输入 Excel 文件
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"错误：Excel 中未找到 sheet: {args.sheet}")
        print(f"可用 sheet: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    ws = wb[args.sheet]
    max_row = ws.max_row
    print(f"Sheet '{args.sheet}' 最大行号: {max_row}")

    # 3. 解析 rows 范围
    try:
        start_row, end_row = parse_rows_spec(args.rows, max_row)
    except Exception as e:
        print(f"错误：解析 rows 参数出错: {e}")
        sys.exit(1)
    data_row_count = end_row - start_row + 1
    print(f"处理行范围: {start_row}-{end_row} (共 {data_row_count} 行)")

    # 4. 加载日志数据
    search_data = load_search_log(search_log_path, args.sheet)
    snapshot_data = load_snapshot_log(snapshot_log_path)

    print(f"加载完成：{len(search_data)} 条搜索记录，{len(snapshot_data)} 条快照记录")

    if not search_data:
        print("警告：没有找到任何搜索记录！请检查：")
        print("1. 搜索日志文件是否存在")
        print("2. 搜索日志中是否包含正确的 sheet 名称")
        print("3. 搜索日志文件格式是否正确")

    if not snapshot_data:
        print("警告：没有找到任何快照记录！请检查：")
        print("1. 快照日志文件是否存在")
        print("2. 快照日志文件格式是否正确")

    # 5. 生成输出文件名并复制文件
    output_filename = generate_output_filename(input_path, data_row_count)
    output_path = os.path.join(base_dir, output_filename)
    shutil.copy2(input_path, output_path)
    print(f"已复制输入文件到: {output_path}")

    # 6. 打开输出文件进行编辑
    wb_out = load_workbook(output_path)
    ws_out = wb_out[args.sheet]

    # 获取最后一列的列号
    last_col = ws_out.max_column
    print(f"输入文件有 {last_col} 列，将添加 14 个新列")

    # 新增列的标题
    new_headers = [
        "原始行号", "关键词", "搜索时间", "搜索耗时", "搜索结果", "搜索错误",
        "快照状态", "快照错误", "链接1", "快照1", "链接2", "快照2", "链接3", "快照3"
    ]

    print(f"正在添加新列标题: {new_headers}")

    # 在第一行添加新列标题
    for i, header in enumerate(new_headers, start=1):
        col_num = last_col + i
        ws_out.cell(row=1, column=col_num, value=header)
        debug_print(f"设置列 {col_num} 为: {header}")

    print(f"列标题添加完成，最终列数: {ws_out.max_column}")

    # 7. 处理每一行数据
    processed_count = 0
    matched_count = 0

    print(f"开始处理 {data_row_count} 行数据...")

    for row_idx in range(start_row, end_row + 1):
        processed_count += 1

        # 生成关键词
        keywords = build_keywords_from_row(ws, row_idx, columns_spec)
        debug_print(f"第 {row_idx} 行关键词: '{keywords}'")

        # 查找匹配的搜索数据
        search_info = search_data.get(keywords)

        if search_info is None:
            print(f"警告：第 {row_idx} 行关键词 '{keywords}' 未找到匹配的搜索记录")
            print(f"  可用的关键词: {list(search_data.keys())[:5]}...")  # 显示前5个可用的关键词
            # 仍然写入原始行号，但其他列留空
            ws_out.cell(row=row_idx, column=last_col + 1, value=row_idx)
            continue

        matched_count += 1
        debug_print(f"第 {row_idx} 行找到匹配记录")

        # 检查搜索结果
        search_error = search_info["search_error"]
        search_result_json = search_info["search_result_json"]

        # 如果搜索成功但结果为空，视为错误
        if not search_error and not search_result_json:
            search_error = "搜索成功但无结果"

        # 计算快照状态
        snapshot_status, snapshot_errors = calculate_snapshot_status(
            search_result_json,
            search_info["urls"],
            snapshot_data
        )

        # 填充数据
        col_offset = last_col + 1
        ws_out.cell(row=row_idx, column=col_offset, value=row_idx)  # 原始行号
        ws_out.cell(row=row_idx, column=col_offset + 1, value=keywords)  # 关键词
        ws_out.cell(row=row_idx, column=col_offset + 2, value=search_info["search_time"])  # 搜索时间
        ws_out.cell(row=row_idx, column=col_offset + 3, value=search_info["search_duration_ms"])  # 搜索耗时
        ws_out.cell(row=row_idx, column=col_offset + 4, value=search_result_json)  # 搜索结果
        ws_out.cell(row=row_idx, column=col_offset + 5, value=search_error)  # 搜索错误
        ws_out.cell(row=row_idx, column=col_offset + 6, value=snapshot_status)  # 快照状态
        ws_out.cell(row=row_idx, column=col_offset + 7, value=snapshot_errors)  # 快照错误

        # 链接和快照列
        for i, url in enumerate(search_info["urls"]):
            if url:
                ws_out.cell(row=row_idx, column=col_offset + 8 + i*2, value=url)  # 链接
                snap_info = snapshot_data.get(url)
                if snap_info and snap_info.get("snapshot_path"):
                    snapshot_path = args.snapshot_prefix + snap_info["snapshot_path"]
                    ws_out.cell(row=row_idx, column=col_offset + 9 + i*2, value=snapshot_path)  # 快照

    # 8. 调整新增列的宽度，使其在Excel中更容易看到
    print("正在调整列宽...")
    for col_num in range(last_col + 1, last_col + 15):  # 新增的14列 + 原始行号列
        col_letter = ws_out.cell(row=1, column=col_num).column_letter
        if col_num == last_col + 1:  # 原始行号列
            ws_out.column_dimensions[col_letter].width = 10
        elif col_num <= last_col + 6:  # 关键词、搜索时间等较短的列
            ws_out.column_dimensions[col_letter].width = 15
        elif col_num == last_col + 7:  # 快照状态
            ws_out.column_dimensions[col_letter].width = 10
        elif col_num == last_col + 8:  # 快照错误
            ws_out.column_dimensions[col_letter].width = 30
        else:  # URL列
            ws_out.column_dimensions[col_letter].width = 50

    # 9. 保存输出文件
    wb_out.save(output_path)
    print(f"输出文件已保存: {output_path}")
    print(f"[成功] 已添加 14 个新列，列宽已自动调整")
    print(f"[提示] 在Excel中如看不到新列，请尝试：")
    print(f"   1. 使用水平滚动条滚动到右侧")
    print(f"   2. 选择所有列，右键→列宽→设置为合适值")
    print(f"   3. 新增列从第{last_col + 1}列开始")

    print("")
    print("=" * 70)
    print("处理完成")
    print(f"  处理数据行数: {data_row_count}")
    print(f"  找到匹配记录: {matched_count}")
    print(f"  未匹配行数: {data_row_count - matched_count}")
    print("=" * 70)


if __name__ == "__main__":
    main()
