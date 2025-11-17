#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
search_snapshot.py

按 Excel 行构造搜索关键字，调用 ScrapingBee Google Search API，
取前 3 条 URL 做截图快照，并将过程记录到 log.csv 中。

依赖：
    pip install openpyxl requests scrapingbee

使用示例：
    python search_snapshot.py \
        --input-file=/path/to/file.xlsx \
        --sheet=Sheet1 \
        --search-columns=C*,D \
        --rows=3+ \
        --debug

配置文件：
    同目录下可选 config.toml，结构示例：

    [scrapingbee]
    timeout_seconds = 120
    concurrency = 3
    retry_times = 1

环境变量：
    当前目录 .env 文件中可设置：
        SCRAPINGBEE_API_KEY=xxxxxx
"""

import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Tuple, Any

import requests
import urllib3
from openpyxl import load_workbook
from scrapingbee import ScrapingBeeClient

# 禁用 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- toml 解析：兼容 3.11+ 的 tomllib 与 3.10 的 tomli ---
try:
    import tomllib  # Python 3.11+
except ImportError:  # pragma: no cover
    try:
        import tomli as tomllib  # 需要 pip install tomli
    except ImportError:
        tomllib = None  # 后面代码会处理这个情况

# ----------------- 全局状态 -----------------

DEBUG = False

# 搜索缓存：keywords -> dict(search_result_json, search_error, urls)
search_cache_lock = threading.Lock()
search_cache: Dict[str, Dict[str, Any]] = {}

# 快照缓存：url -> dict(snapshot_path, sheets, rows, keywords, error, time, is_direct, size)
snapshot_cache_lock = threading.Lock()
snapshot_cache: Dict[str, Dict[str, Any]] = {}

# 已完成行： (sheet_name, row_number)
row_done_lock = threading.Lock()
row_done_set: set[Tuple[str, int]] = set()

# 进度统计
progress_lock = threading.Lock()
total_tasks = 0
finished_tasks = 0
success_tasks = 0
failed_tasks = 0


# ----------------- 工具函数 -----------------

# 任务上下文：用于在并行任务中传递任务信息
class TaskContext:
    def __init__(self, worker_id: int, task_index: int, total_tasks: int):
        self.worker_id = worker_id
        self.task_index = task_index
        self.total_tasks = total_tasks
    
    def prefix(self) -> str:
        """返回任务前缀"""
        return f"#{self.worker_id} - [{self.task_index}/{self.total_tasks}]"


def debug_print(*args, task_ctx: Optional[TaskContext] = None, **kwargs):
    if DEBUG:
        now = dt.datetime.now().strftime("%H:%M:%S")
        prefix = f"{task_ctx.prefix()} " if task_ctx else ""
        print(f"[{now}] [DEBUG] {prefix}", *args, **kwargs)


def log_print(*args, task_ctx: Optional[TaskContext] = None, **kwargs):
    """普通控制台输出，带时间前缀"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    prefix = f"{task_ctx.prefix()} " if task_ctx else ""
    print(f"[{now}] {prefix}", *args, **kwargs)


def info_print(*args, task_ctx: Optional[TaskContext] = None, **kwargs):
    """信息输出"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    prefix = f"{task_ctx.prefix()} " if task_ctx else ""
    print(f"[{now}] [INFO] {prefix}", *args, **kwargs)


def error_print(*args, task_ctx: Optional[TaskContext] = None, **kwargs):
    """错误输出"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    prefix = f"{task_ctx.prefix()} " if task_ctx else ""
    print(f"[{now}] [ERROR] {prefix}", *args, **kwargs)


def progress_print(current: int, total: int, success: int, failed: int, task_ctx: Optional[TaskContext] = None):
    """进度输出"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    percentage = (current / total * 100) if total > 0 else 0
    prefix = f"{task_ctx.prefix()} " if task_ctx else ""
    print(f"[{now}] [进度] {prefix}{current}/{total} ({percentage:.1f}%) | 成功: {success} | 失败: {failed}")


def load_env_file(path: str = ".env"):
    """简单解析 .env 文件，将 KEY=VALUE 写入 os.environ（若不存在则跳过）"""
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip()
                # 去掉包围的引号
                if (value.startswith('"') and value.endswith('"')) or (
                    value.startswith("'") and value.endswith("'")
                ):
                    value = value[1:-1]
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception as e:
        log_print("加载 .env 文件出错：", e)


def load_config(config_path: str) -> Dict[str, Any]:
    """读取 config.toml 中 [scrapingbee] 配置，不存在则使用默认值"""
    default_cfg = {
        "timeout_seconds": 120,
        "concurrency": 1,
        "retry_times": 1,
        "proxy": None,
    }
    if not os.path.exists(config_path):
        debug_print("config.toml 不存在，使用默认配置。")
        return default_cfg

    if tomllib is None:
        log_print("警告：未安装 tomllib/tomli，无法解析 config.toml，使用默认配置。")
        return default_cfg

    try:
        with open(config_path, "rb") as f:
            data = tomllib.load(f)
        bee_cfg = data.get("scrapingbee", {}) or {}
        for key in default_cfg:
            if key in bee_cfg:
                default_cfg[key] = bee_cfg[key]
        debug_print("加载 config.toml 成功：", default_cfg)
        return default_cfg
    except Exception as e:
        log_print("解析 config.toml 出错，使用默认配置。错误：", e)
        return default_cfg


def column_letters_to_index(letters: str) -> int:
    """将 Excel 列字母（如 'A', 'C', 'AA'）转换为 1-based 列索引"""
    letters = letters.upper()
    result = 0
    for ch in letters:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"非法列字母: {letters}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def parse_search_columns(spec: str) -> List[Tuple[int, bool]]:
    """
    解析 --search-columns 参数，如 "C*,D,AA*"
    返回列表 [(col_index, exact_match), ...]
    exact_match=True 表示需要加双引号
    """
    result: List[Tuple[int, bool]] = []
    for token in spec.split(","):
        token = token.strip()
        if not token:
            continue
        exact = token.endswith("*")
        col_letters = token[:-1] if exact else token
        col_index = column_letters_to_index(col_letters)
        result.append((col_index, exact))
    if not result:
        raise ValueError("search-columns 解析结果为空，请检查参数。")
    return result


def parse_rows_spec(spec: str, max_row: int) -> Tuple[int, int]:
    """
    解析 --rows 参数：
        '3+'  -> (3, max_row)
        '3-9' -> (3, 9)
    返回 (start_row, end_row)，均为闭区间
    """
    spec = spec.strip()
    if spec.endswith("+"):
        start = spec[:-1]
        if not start.isdigit():
            raise ValueError(f"rows 参数非法：{spec}")
        start_row = int(start)
        end_row = max_row
    else:
        if "-" not in spec:
            raise ValueError(f"rows 参数非法：{spec}")
        parts = spec.split("-", 1)
        if not (parts[0].isdigit() and parts[1].isdigit()):
            raise ValueError(f"rows 参数非法：{spec}")
        start_row = int(parts[0])
        end_row = int(parts[1])
    if start_row < 1 or end_row < start_row:
        raise ValueError(f"rows 范围非法：{start_row}-{end_row}")
    # 同时不允许超过 max_row
    if start_row > max_row:
        raise ValueError(f"rows 起始行 {start_row} 大于 sheet 最大行 {max_row}")
    end_row = min(end_row, max_row)
    return start_row, end_row


def sha1_hex(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


# ----------------- ScrapingBee 调用 -----------------

def search_google_once(api_key: str, keywords: str, language: str, timeout: int, proxies: Optional[Dict[str, str]] = None) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str], float]:
    """
    调用一次 ScrapingBee Google Search。
    返回 (organic_results, error_message, duration_seconds)
    error_message 为 None 表示成功。
    """
    url = "https://app.scrapingbee.com/api/v1/google"
    params = {
        "api_key": api_key,
        "search": keywords,
        "language": language,
    }
    start = time.monotonic()
    try:
        response = requests.get(url, params=params, timeout=timeout, proxies=proxies)
        duration = time.monotonic() - start
        status = response.status_code
        if status != 200:
            try:
                txt = response.text[:200]
            except Exception:
                txt = ""
            return None, f"HTTP {status}: {txt}", duration
        data = response.json()
        organic = data.get("organic_results", [])
        if not isinstance(organic, list):
            return None, "organic_results 非列表", duration
        return organic, None, duration
    except Exception as e:
        duration = time.monotonic() - start
        return None, f"请求异常: {e}", duration


def search_google_with_retry(api_key: str, keywords: str, language: str, timeout: int, retry_times: int, proxies: Optional[Dict[str, str]] = None, task_ctx: Optional[TaskContext] = None) -> Tuple[List[Dict[str, Any]], str, float]:
    """
    带重试的搜索。
    返回 (organic_results_list, error_message, total_duration)
    error_message 为空字符串表示无错误。
    """
    total_duration = 0.0
    last_error = ""
    attempts = 1 + max(retry_times, 0)
    for attempt in range(1, attempts + 1):
        organic, err, dur = search_google_once(api_key, keywords, language, timeout, proxies)
        total_duration += dur
        if err is None:
            log_print(f"[搜索] ✓ 成功 | 关键字: {keywords} | 耗时: {dur:.2f}s", task_ctx=task_ctx)
        else:
            log_print(f"[搜索] ✗ 失败 (尝试 {attempt}/{attempts}) | 关键字: {keywords} | 耗时: {dur:.2f}s | 错误: {err}", task_ctx=task_ctx)
        if err is None:
            return organic, "", total_duration
        last_error = err
        # 简单退避
        time.sleep(min(1.0 * attempt, 5.0))
    return [], last_error, total_duration


def screenshot_once(client: ScrapingBeeClient, url: str, save_full_path: str, timeout: int, task_ctx: Optional[TaskContext] = None) -> Optional[str]:
    """
    截图一次；成功返回 None (表示无错误)，失败返回 error_message。
    """
    start = time.monotonic()
    try:
        os.makedirs(os.path.dirname(save_full_path), exist_ok=True)
        response = client.get(
            url,
            params={
                "screenshot": True,
                "screenshot_full_page": True,
            },
            timeout=timeout,
        )
        duration = time.monotonic() - start
        status = getattr(response, 'status_code', 'N/A')
        file_size = len(getattr(response, "content", b""))
        log_print(f"[截图] ✓ 成功 | URL: {url} | 状态: {status} | 大小: {file_size/1024:.1f}KB | 耗时: {duration:.2f}s", task_ctx=task_ctx)
        content = getattr(response, "content", None)
        if not content:
            return f"empty content for url={url}"
        with open(save_full_path, "wb") as f:
            f.write(content)
        return None
    except Exception as e:
        duration = time.monotonic() - start
        log_print(f"[截图] ✗ 失败 | URL: {url} | 耗时: {duration:.2f}s | 错误: {str(e)}", task_ctx=task_ctx)
        return f"exception for url={url}: {e}"


def screenshot_with_retry(client: ScrapingBeeClient, url: str, save_full_path: str, timeout: int, retry_times: int, task_ctx: Optional[TaskContext] = None) -> Tuple[bool, List[str]]:
    """
    带重试的截图。
    返回 (success, error_messages_list)
    """
    errors: List[str] = []
    attempts = 1 + max(retry_times, 0)
    for attempt in range(1, attempts + 1):
        err = screenshot_once(client, url, save_full_path, timeout, task_ctx)
        if err is None:
            return True, []
        errors.append(f"attempt {attempt}: {err}")
        time.sleep(min(1.0 * attempt, 5.0))
    return False, errors


def is_direct_downloadable(url: str) -> bool:
    """判断 URL 是否可直接下载（根据扩展名）"""
    try:
        from urllib.parse import urlparse, unquote
        parsed = urlparse(url)
        path = unquote(parsed.path).lower()
        return any(path.endswith(ext) for ext in DIRECT_DOWNLOAD_EXTENSIONS)
    except Exception:
        return False


def direct_download_once(url: str, save_full_path: str, timeout: int, proxies: Optional[Dict[str, str]] = None, task_ctx: Optional[TaskContext] = None) -> Tuple[Optional[int], Optional[str]]:
    """
    直接下载文件。
    返回 (file_size_bytes, error_message)
    error_message 为 None 表示成功。
    """
    start = time.monotonic()
    try:
        os.makedirs(os.path.dirname(save_full_path), exist_ok=True)
        
        # 设置浏览器请求头，避免 406 错误
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': '*/*',  # 接受所有内容类型
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        
        # 关闭 SSL 验证以避免证书问题，允许重定向
        response = requests.get(url, headers=headers, timeout=timeout, stream=True, verify=False, proxies=proxies, allow_redirects=True)
        duration = time.monotonic() - start
        
        if response.status_code != 200:
            log_print(f"[下载] ✗ 失败 | URL: {url} | 状态: {response.status_code} | 耗时: {duration:.2f}s", task_ctx=task_ctx)
            return None, f"HTTP {response.status_code}"
        
        # 写入文件
        file_size = 0
        with open(save_full_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    file_size += len(chunk)
        
        duration = time.monotonic() - start
        log_print(f"[下载] ✓ 成功 | URL: {url} | 大小: {file_size/1024:.1f}KB | 耗时: {duration:.2f}s", task_ctx=task_ctx)
        return file_size, None
    except Exception as e:
        duration = time.monotonic() - start
        log_print(f"[下载] ✗ 失败 | URL: {url} | 耗时: {duration:.2f}s | 错误: {str(e)}", task_ctx=task_ctx)
        return None, f"下载异常: {e}"


def direct_download_with_retry(url: str, save_full_path: str, timeout: int, retry_times: int, proxies: Optional[Dict[str, str]] = None, task_ctx: Optional[TaskContext] = None) -> Tuple[bool, Optional[int], List[str]]:
    """
    带重试的直接下载。
    返回 (success, file_size_bytes, error_messages_list)
    """
    errors: List[str] = []
    attempts = 1 + max(retry_times, 0)
    for attempt in range(1, attempts + 1):
        file_size, err = direct_download_once(url, save_full_path, timeout, proxies, task_ctx)
        if err is None:
            return True, file_size, []
        errors.append(f"attempt {attempt}: {err}")
        time.sleep(min(1.0 * attempt, 5.0))
    return False, None, errors


# ----------------- 日志文件相关 -----------------

SEARCH_LOG_HEADER = [
    "sheet",
    "row",
    "keywords",
    "search_time",
    "search_duration_ms",
    "search_result_json",
    "search_error",
    "url1",
    "url2",
    "url3",
]

SNAPSHOT_LOG_HEADER = [
    "url",
    "sheets",
    "rows",
    "keywords",
    "snapshot_path",
    "snapshot_error",
    "snapshot_time",
    "is_direct_download",
    "file_size_bytes",
]

# 可直接下载的文件扩展名
DIRECT_DOWNLOAD_EXTENSIONS = {
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx',
    '.zip', '.png', '.jpg', '.jpeg', '.gif'
}


def ensure_log_header(log_path: str, header: List[str]):
    """如果日志文件不存在，则创建并写入表头"""
    if not os.path.exists(log_path):
        with open(log_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(header)
        debug_print(f"创建日志文件并写入表头: {log_path}")


def load_existing_logs(search_log_path: str, snapshot_log_path: str, snapshot_root: str):
    """
    启动时读取已有的 search.csv 和 snapshot.csv，更新：
        - search_cache (按 keywords)
        - snapshot_cache (按 url)
        - row_done_set
    """
    # 1. 读取 search.csv，恢复搜索缓存
    search_rows: Dict[Tuple[str, int], Dict[str, Any]] = {}
    if os.path.exists(search_log_path):
        with open(search_log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                sheet_name = row.get("sheet", "")
                row_no_str = row.get("row", "")
                keywords = (row.get("keywords") or "").strip()
                search_result_json = row.get("search_result_json") or ""
                search_error = row.get("search_error") or ""
                url1 = row.get("url1") or ""
                url2 = row.get("url2") or ""
                url3 = row.get("url3") or ""
                
                try:
                    row_no = int(row_no_str)
                except ValueError:
                    continue
                
                # 保存该行信息，用于后续判断完成状态
                search_rows[(sheet_name, row_no)] = {
                    "keywords": keywords,
                    "search_error": search_error,
                    "urls": [url1, url2, url3],
                }
                
                # 恢复搜索缓存
                if keywords and search_result_json:
                    with search_cache_lock:
                        if keywords not in search_cache:
                            try:
                                parsed = json.loads(search_result_json)
                            except Exception:
                                parsed = []
                            search_cache[keywords] = {
                                "results": parsed,
                                "search_error": search_error,
                            }
        debug_print(f"从 search.csv 恢复 {len(search_rows)} 行，{len(search_cache)} 个搜索缓存")
    else:
        debug_print("search.csv 不存在，无需恢复搜索状态。")
    
    # 2. 读取 snapshot.csv，恢复快照缓存
    if os.path.exists(snapshot_log_path):
        with open(snapshot_log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = row.get("url", "")
                sheets_str = row.get("sheets", "")
                rows_str = row.get("rows", "")
                keywords_str = row.get("keywords", "")
                snapshot_path = row.get("snapshot_path", "")
                snapshot_error = row.get("snapshot_error", "")
                snapshot_time = row.get("snapshot_time", "")
                is_direct = row.get("is_direct_download", "")
                file_size_str = row.get("file_size_bytes", "")
                
                if not url:
                    continue
                
                # 验证文件是否存在
                if snapshot_path and not snapshot_error:
                    full_path = os.path.join(snapshot_root, snapshot_path)
                    if os.path.exists(full_path):
                        with snapshot_cache_lock:
                            if url not in snapshot_cache:
                                try:
                                    file_size = int(file_size_str) if file_size_str else 0
                                except ValueError:
                                    file_size = 0
                                
                                snapshot_cache[url] = {
                                    "snapshot_path": snapshot_path,
                                    "sheets": sheets_str.split("\n") if sheets_str else [],
                                    "rows": rows_str.split("\n") if rows_str else [],
                                    "keywords": keywords_str.split("\n") if keywords_str else [],
                                    "snapshot_error": snapshot_error,
                                    "snapshot_time": snapshot_time,
                                    "is_direct_download": is_direct.lower() == "true",
                                    "file_size_bytes": file_size,
                                }
        debug_print(f"从 snapshot.csv 恢复 {len(snapshot_cache)} 个快照缓存")
    else:
        debug_print("snapshot.csv 不存在，无需恢复快照状态。")
    
    # 3. 判断行完成状态
    for (sheet_name, row_no), search_info in search_rows.items():
        if search_info["search_error"]:
            # 搜索失败，不算完成
            continue
        
        urls = [u for u in search_info["urls"] if u]
        if not urls:
            # 没有 URL，算完成
            row_done_set.add((sheet_name, row_no))
            continue
        
        # 检查所有 URL 是否都有成功的快照
        all_done = True
        with snapshot_cache_lock:
            for url in urls:
                cached = snapshot_cache.get(url)
                if not cached or cached.get("snapshot_error"):
                    all_done = False
                    break
        
        if all_done:
            row_done_set.add((sheet_name, row_no))
    
    debug_print(f"恢复状态完成：{len(row_done_set)} 行已完成")


# ----------------- 核心 worker -----------------

def build_keywords_from_row(ws, row_idx: int, columns_spec: List[Tuple[int, bool]]) -> str:
    """
    从给定行构造搜索关键字字符串。
    columns_spec: [(col_index, exact), ...]
    """
    parts: List[str] = []
    for col_index, exact in columns_spec:
        cell = ws.cell(row=row_idx, column=col_index)
        value = cell.value
        if value is None:
            continue
        s = str(value)
        s = s.replace("\n", " ").strip()
        if not s:
            continue
        if exact:
            parts.append(f'"{s}"')
        else:
            parts.append(s)
    keywords = " ".join(parts).strip()
    return keywords


def process_row_task(
    sheet_name: str,
    row_idx: int,
    ws,
    columns_spec: List[Tuple[int, bool]],
    api_key: str,
    cfg: Dict[str, Any],
    snapshot_root: str,
    search_log_writer: csv.writer,
    search_log_file,
    search_log_lock: threading.Lock,
    snapshot_log_writer: csv.writer,
    snapshot_log_file,
    snapshot_log_lock: threading.Lock,
    task_ctx: Optional[TaskContext] = None,
):
    global finished_tasks, success_tasks, failed_tasks

    # 1. 如果行已完成，直接跳过
    with row_done_lock:
        if (sheet_name, row_idx) in row_done_set:
            debug_print(f"行已完成，跳过: {sheet_name}#{row_idx}", task_ctx=task_ctx)
            return

    # 2. 构造关键字
    keywords = build_keywords_from_row(ws, row_idx, columns_spec)
    keywords = keywords.strip()

    if not keywords:
        msg = f"Sheet={sheet_name} Row={row_idx} 搜索列全为空，跳过此行。"
        log_print(msg, task_ctx=task_ctx)
        search_time_iso = dt.datetime.now().astimezone().isoformat()
        # 写入 search.csv
        with search_log_lock:
            search_log_writer.writerow([
                sheet_name,
                row_idx,
                keywords,
                search_time_iso,
                0,  # search_duration_ms
                "",  # search_result_json
                "empty keywords",  # search_error
                "",  # url1
                "",  # url2
                "",  # url3
            ])
            search_log_file.flush()

        with progress_lock:
            finished_tasks += 1
            success_tasks += 1
            progress_print(finished_tasks, total_tasks, success_tasks, failed_tasks, task_ctx=task_ctx)
        
        # 标记为完成
        with row_done_lock:
            row_done_set.add((sheet_name, row_idx))
        return

    # 3. 搜索（使用缓存 + 重试）
    timeout = int(cfg["timeout_seconds"])
    retry_times = int(cfg["retry_times"])
    language = "en"
    proxy_url = cfg.get("proxy")
    proxies = None
    if proxy_url:
        proxies = {
            "http": proxy_url,
            "https": proxy_url,
        }

    # 搜索缓存检查
    with search_cache_lock:
        cached = search_cache.get(keywords)

    if cached is not None:
        organic_results = cached["results"]
        search_error = cached.get("search_error", "")
        search_duration_ms = 0
        search_time_iso = dt.datetime.now().astimezone().isoformat()
        debug_print(f"使用已有搜索缓存：keywords={keywords!r}", task_ctx=task_ctx)
    else:
        search_time_iso = dt.datetime.now().astimezone().isoformat()
        organic_results, search_error, dur = search_google_with_retry(
            api_key=api_key,
            keywords=keywords,
            language=language,
            timeout=timeout,
            retry_times=retry_times,
            proxies=proxies,
            task_ctx=task_ctx,
        )
        search_duration_ms = int(dur * 1000)
        # 写入缓存
        with search_cache_lock:
            search_cache[keywords] = {
                "results": organic_results,
                "search_error": search_error,
            }

    # 转成 JSON 保存
    try:
        search_result_json = json.dumps(organic_results, ensure_ascii=False)
    except Exception:
        search_result_json = ""

    # 4. 取前 3 条 URL
    urls: List[str] = []
    for item in organic_results[:3]:
        url = item.get("url") if isinstance(item, dict) else None
        if url:
            urls.append(str(url))
    while len(urls) < 3:
        urls.append("")
    
    # 5. 写入 search.csv
    with search_log_lock:
        search_log_writer.writerow([
            sheet_name,
            row_idx,
            keywords,
            search_time_iso,
            search_duration_ms,
            search_result_json,
            search_error,
            urls[0],
            urls[1],
            urls[2],
        ])
        search_log_file.flush()

    # 6. 处理快照（如果搜索失败，则跳过）
    has_errors = False
    if search_error:
        log_print(f"[跳过] 搜索失败，不进行快照 | 关键字: {keywords} | 错误: {search_error}", task_ctx=task_ctx)
        has_errors = True
    else:
        # 创建 ScrapingBee 客户端（代理已通过环境变量设置）
        client = ScrapingBeeClient(api_key=api_key)
        
        for url in urls:
            if not url:
                continue
            
            # 检查快照缓存
            with snapshot_cache_lock:
                cached = snapshot_cache.get(url)
            
            if cached and not cached.get("snapshot_error"):
                # 已有成功的快照，追加 sheets/rows/keywords
                debug_print(f"使用已有快照：url={url}")
                
                # 验证文件存在
                snapshot_path = cached["snapshot_path"]
                full_path = os.path.join(snapshot_root, snapshot_path)
                if os.path.exists(full_path):
                    # 追加当前的 sheet/row/keywords（去重）
                    sheets_set = set(cached["sheets"])
                    rows_set = set(cached["rows"])
                    keywords_set = set(cached["keywords"])
                    
                    sheets_set.add(sheet_name)
                    rows_set.add(str(row_idx))
                    keywords_set.add(keywords)
                    
                    # 更新缓存
                    with snapshot_cache_lock:
                        snapshot_cache[url]["sheets"] = sorted(sheets_set)
                        snapshot_cache[url]["rows"] = sorted(rows_set, key=lambda x: int(x) if x.isdigit() else 0)
                        snapshot_cache[url]["keywords"] = sorted(keywords_set)
                    
                    # 重新写入整个 snapshot.csv（需要读取所有记录并更新）
                    # 注：为了简化，我们只追加新记录或更新现有记录
                    # 这里我们采用追加方式，后续可以优化为定期合并去重
                    continue
                else:
                    # 文件不存在，缓存作废，需要重新下载
                    with snapshot_cache_lock:
                        snapshot_cache.pop(url, None)
            
            # 需要新建快照
            h = sha1_hex(url)
            is_direct = is_direct_downloadable(url)
            
            # 根据文件类型选择扩展名
            if is_direct:
                # 尝试从 URL 中提取扩展名
                from urllib.parse import urlparse, unquote
                try:
                    parsed = urlparse(url)
                    path = unquote(parsed.path)
                    ext = os.path.splitext(path)[1].lower()
                    if not ext or ext not in DIRECT_DOWNLOAD_EXTENSIONS:
                        ext = ".pdf"  # 默认
                except Exception:
                    ext = ".pdf"
            else:
                ext = ".png"
            
            rel_path = os.path.join(h[:2], h[2:4], h[4:] + ext)
            full_path = os.path.join(snapshot_root, rel_path)
            
            snapshot_time = dt.datetime.now().astimezone().isoformat()
            
            if is_direct:
                # 直接下载
                ok, file_size, errs = direct_download_with_retry(
                    url=url,
                    save_full_path=full_path,
                    timeout=timeout,
                    retry_times=retry_times,
                    proxies=proxies,
                    task_ctx=task_ctx,
                )
            else:
                # 使用 ScrapingBee 截图
                ok, errs = screenshot_with_retry(
                    client=client,
                    url=url,
                    save_full_path=full_path,
                    timeout=timeout,
                    retry_times=retry_times,
                    task_ctx=task_ctx,
                )
                if ok:
                    # 获取文件大小
                    try:
                        file_size = os.path.getsize(full_path)
                    except Exception:
                        file_size = 0
                else:
                    file_size = 0
            
            if ok:
                snapshot_error = ""
            else:
                snapshot_error = "; ".join(errs)
                has_errors = True
            
            # 写入 snapshot.csv
            with snapshot_log_lock:
                snapshot_log_writer.writerow([
                    url,
                    sheet_name,
                    str(row_idx),
                    keywords,
                    rel_path if ok else "",
                    snapshot_error,
                    snapshot_time if ok else "",
                    "true" if is_direct else "false",
                    str(file_size) if ok else "",
                ])
                snapshot_log_file.flush()
            
            # 更新缓存
            if ok:
                with snapshot_cache_lock:
                    snapshot_cache[url] = {
                        "snapshot_path": rel_path,
                        "sheets": [sheet_name],
                        "rows": [str(row_idx)],
                        "keywords": [keywords],
                        "snapshot_error": "",
                        "snapshot_time": snapshot_time,
                        "is_direct_download": is_direct,
                        "file_size_bytes": file_size,
                    }

    # 7. 更新进度与完成标记
    with progress_lock:
        finished_tasks += 1
        if not has_errors:
            success_tasks += 1
        else:
            failed_tasks += 1
        progress_print(finished_tasks, total_tasks, success_tasks, failed_tasks, task_ctx=task_ctx)

    # 将已完成的行加入 row_done_set（仅在无错误时）
    if not has_errors:
        with row_done_lock:
            row_done_set.add((sheet_name, row_idx))


# ----------------- 主函数 -----------------

def main():
    global DEBUG, total_tasks

    parser = argparse.ArgumentParser(description="Excel 驱动的 ScrapingBee 搜索+快照脚本")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet", required=True, help="Sheet 名称（不是序号）")
    parser.add_argument("--search-columns", required=True, help="搜索列设置，例如 C*,D")
    parser.add_argument("--rows", required=True, help="行范围，例如 3+ 或 3-9")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    args = parser.parse_args()

    DEBUG = bool(args.debug)

    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path):
        error_print(f"输入文件不存在: {input_path}")
        sys.exit(1)

    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]

    snapshot_root = os.path.join(base_dir, f"{base_name}-snapshot")
    search_log_path = os.path.join(base_dir, f"{base_name}.search.csv")
    snapshot_log_path = os.path.join(base_dir, f"{base_name}.snapshot.csv")
    
    # 配置文件查找顺序：1. 脚本所在目录 2. 输入文件所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config.toml")
    if not os.path.exists(config_path):
        config_path = os.path.join(base_dir, "config.toml")

    info_print("=" * 70)
    info_print("配置信息")
    info_print(f"  输入文件: {input_path}")
    info_print(f"  快照目录: {snapshot_root}")
    info_print(f"  搜索日志: {search_log_path}")
    info_print(f"  快照日志: {snapshot_log_path}")
    info_print(f"  配置文件: {config_path}")
    info_print("=" * 70)

    # 1. 加载 .env 和配置
    load_env_file(".env")
    cfg = load_config(config_path)
    
    proxy_url = cfg.get("proxy")
    if proxy_url:
        # 设置环境变量，使 ScrapingBee 客户端也能使用代理
        os.environ["HTTP_PROXY"] = proxy_url
        os.environ["HTTPS_PROXY"] = proxy_url
        info_print(f"代理设置: {proxy_url}")
    else:
        debug_print("未配置代理")

    api_key = os.environ.get("SCRAPINGBEE_API_KEY", "").strip()
    if not api_key:
        error_print("未找到 SCRAPINGBEE_API_KEY，请在 .env 或环境变量中设置。")
        sys.exit(1)

    # 2. 解析 search-columns
    try:
        columns_spec = parse_search_columns(args.search_columns)
        debug_print("columns_spec =", columns_spec)
    except Exception as e:
        error_print(f"解析 search-columns 出错: {e}")
        sys.exit(1)

    # 3. 打开 Excel
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        error_print(f"Excel 中未找到 sheet: {args.sheet}")
        info_print(f"可用 sheet: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    ws = wb[args.sheet]
    max_row = ws.max_row
    info_print(f"Sheet '{args.sheet}' 最大行号: {max_row}")

    # 4. 解析 rows 范围
    try:
        start_row, end_row = parse_rows_spec(args.rows, max_row)
    except Exception as e:
        error_print(f"解析 rows 参数出错: {e}")
        sys.exit(1)
    info_print(f"处理行范围: {start_row}-{end_row}")

    # 5. 准备日志文件 & 恢复状态
    ensure_log_header(search_log_path, SEARCH_LOG_HEADER)
    ensure_log_header(snapshot_log_path, SNAPSHOT_LOG_HEADER)
    load_existing_logs(search_log_path, snapshot_log_path, snapshot_root)

    # 6. 构建任务列表（仅未完成的行）
    tasks: List[int] = []
    for row_idx in range(start_row, end_row + 1):
        with row_done_lock:
            if (args.sheet, row_idx) in row_done_set:
                debug_print(f"行已完成，略过: {args.sheet}#{row_idx}")
                continue
        tasks.append(row_idx)

    total_tasks = len(tasks)
    if total_tasks == 0:
        info_print("指定范围内的行已全部处理完成，无任务可执行。")
        return

    info_print("=" * 70)
    info_print(f"开始处理 | 总任务数: {total_tasks}")
    info_print("=" * 70)

    # 7. 打开日志文件（append 模式），初始化 writer
    search_log_lock = threading.Lock()
    snapshot_log_lock = threading.Lock()
    
    with open(search_log_path, "a", newline="", encoding="utf-8") as search_log_file, \
         open(snapshot_log_path, "a", newline="", encoding="utf-8") as snapshot_log_file:
        
        search_log_writer = csv.writer(search_log_file)
        snapshot_log_writer = csv.writer(snapshot_log_file)

        # 8. 并发执行
        concurrency = int(cfg["concurrency"])
        if concurrency < 1:
            concurrency = 1
        info_print(f"并发线程数: {concurrency}")
        info_print("")

        # 为每个任务创建任务上下文
        task_contexts = {}
        for idx, row_idx in enumerate(tasks, start=1):
            # worker_id 会在任务执行时动态分配（使用线程ID）
            # 这里先创建上下文，worker_id 会在任务开始时设置
            task_contexts[row_idx] = TaskContext(worker_id=0, task_index=idx, total_tasks=total_tasks)
        
        # 用于分配 worker_id 的计数器
        worker_id_counter = 0
        worker_id_lock = threading.Lock()
        
        def create_task_with_context(row_idx):
            """创建带任务上下文的任务包装函数"""
            nonlocal worker_id_counter
            # 获取 worker_id
            with worker_id_lock:
                worker_id_counter += 1
                worker_id = worker_id_counter
            
            # 更新任务上下文的 worker_id
            task_ctx = task_contexts[row_idx]
            task_ctx.worker_id = worker_id
            
            return process_row_task(
                args.sheet,
                row_idx,
                ws,
                columns_spec,
                api_key,
                cfg,
                snapshot_root,
                search_log_writer,
                search_log_file,
                search_log_lock,
                snapshot_log_writer,
                snapshot_log_file,
                snapshot_log_lock,
                task_ctx=task_ctx,
            )
        
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            future_to_row = {
                executor.submit(create_task_with_context, row_idx): row_idx
                for row_idx in tasks
            }

            # 等待所有任务完成
            for future in as_completed(future_to_row):
                row_idx = future_to_row[future]
                try:
                    future.result()
                except Exception as e:
                    task_ctx = task_contexts.get(row_idx)
                    error_print(f"行 {row_idx} 处理过程中出现未捕获异常: {e}", task_ctx=task_ctx)

    info_print("")
    info_print("=" * 70)
    info_print("任务完成")
    info_print(f"  总任务: {total_tasks}")
    info_print(f"  成功: {success_tasks}")
    info_print(f"  失败: {failed_tasks}")
    info_print("=" * 70)


if __name__ == "__main__":
    main()
