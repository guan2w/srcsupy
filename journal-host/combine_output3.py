#!/usr/bin/env python3
"""
三种方法对比整合工具 - combine_output3.py

整合三种提取方式的结果到单表，便于准确率对比评估：
1. 按关键词规则提取（snapshot → extract）
2. AI精准操作提示（url_scan）
3. AI核心目标提示（search）
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("[ERROR] pandas/openpyxl not installed. Run: pip install pandas openpyxl", file=sys.stderr)
    sys.exit(1)


# ========== 关键词列表（显示顺序和格式） ==========

KEYWORD_COLUMNS = [
    "On behalf of",
    "journal of",
    "publication of",
    "Affiliate",
    "Edited",
    "Own",
    "In association with",
    "responsible",
    "supervise",
    "sponsor",
    "Patronage",
    "Compile",
    "partnership",
    "Societies",
    "In cooperation with",
    "The backing of",
    "administrated",
    "Copyright ©",
    "©",
    "Press",
    "Funded"
]


# ========== 工具函数 ==========

def sha1_hex(text: str) -> str:
    """计算字符串的 SHA1 hash"""
    return hashlib.sha1(text.encode("utf-8")).hexdigest()


def parse_rows_range(rows_str: str) -> Tuple[int, Optional[int]]:
    """
    解析行范围字符串
    
    Args:
        rows_str: 行范围，如 "2-99" 或 "2+"
    
    Returns:
        (start_row, end_row)
        - "2+" -> (2, None) 表示从第2行开始，直到空行
        - "2-99" -> (2, 99) 表示第2行到第99行
    """
    rows_str = rows_str.strip()
    
    # 处理 "2+" 格式
    if rows_str.endswith('+'):
        start_row = int(rows_str[:-1])
        return start_row, None
    
    # 处理 "2-99" 格式
    match = re.match(r'(\d+)-(\d+)', rows_str)
    if match:
        start_row = int(match.group(1))
        end_row = int(match.group(2))
        return start_row, end_row
    
    raise ValueError(f"Invalid rows format: {rows_str}. Use '2-99' or '2+'")


# ========== 数据读取 ==========

def read_input_excel(
    excel_path: Path,
    sheet_name: Any,
    start_row: int,
    end_row: Optional[int]
) -> Tuple[List[str], pd.DataFrame]:
    """
    从 Excel 文件读取原始数据
    
    Returns:
        (header_names, dataframe)
    """
    try:
        # 先读取表头（第一行）
        header_df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name,
            nrows=1,
            header=None,
            engine='openpyxl'
        )
        header_names = header_df.iloc[0].tolist()
        
        # 读取数据（从 start_row 开始）
        skiprows = start_row - 1
        
        if end_row is not None:
            nrows = end_row - start_row + 1
        else:
            nrows = None  # 读取到最后
        
        df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name,
            skiprows=skiprows,
            nrows=nrows,
            header=None,
            engine='openpyxl'
        )
        
        # 如果是 "2+" 格式，遇到空行停止
        if end_row is None:
            # 找到第一个空行（所有列都为空）
            for i in range(len(df)):
                if df.iloc[i].isna().all():
                    # 截取到空行之前
                    df = df.iloc[:i]
                    break
        
        # 设置列名为表头
        df.columns = header_names[:len(df.columns)]
        
        return header_names, df
    
    except Exception as e:
        print(f"[ERROR] Failed to read Excel: {e}", file=sys.stderr)
        sys.exit(1)


def get_hash_path(snapshot_dir: Path, url_hash: str) -> Path:
    """获取 hash 分层目录路径"""
    return snapshot_dir / url_hash[:2] / url_hash[2:4] / url_hash[4:]


def load_host_json(snapshot_dir: Path, url_hash: str) -> Optional[List[Dict[str, Any]]]:
    """
    加载 host.json 文件（优先 langextract，回退 regexp）
    
    Returns:
        host_institutions 列表，如果不存在返回 None
    """
    hash_path = get_hash_path(snapshot_dir, url_hash)
    
    # 按优先级尝试读取
    json_files = [
        hash_path / "host-langextract.json",
        hash_path / "host-regexp.json",
    ]
    
    for json_file in json_files:
        if json_file.exists():
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('host_institutions', [])
            except Exception as e:
                print(f"[WARNING] Failed to load {json_file}: {e}", file=sys.stderr)
                continue
    
    return None


def load_extract_results(snapshot_dir: Path, urls: List[str]) -> Tuple[Dict[str, List[str]], List[Dict[str, Any]]]:
    """
    从 snapshot 目录加载方式1的提取结果
    
    Args:
        snapshot_dir: 快照目录
        urls: URL 列表
    
    Returns:
        (keyword_results, unmatched_results)
        - keyword_results: {关键词(小写): [机构名称列表]}
        - unmatched_results: [{keyword: str, name: str}, ...] 未匹配任何关键词的项
    """
    keyword_results = {}
    unmatched_results = []
    
    # 获取所有关键词的小写集合
    keyword_lower_set = {kw.lower() for kw in KEYWORD_COLUMNS}
    
    for url in urls:
        if not url:
            continue
        
        url_hash = sha1_hex(url)
        institutions = load_host_json(snapshot_dir, url_hash)
        
        if institutions:
            for inst in institutions:
                keyword = inst.get('matched_keyword', '').strip().lower()
                name = inst.get('name', '').strip()
                
                if not name:
                    continue
                
                if keyword and keyword in keyword_lower_set:
                    # 匹配到关键词
                    if keyword not in keyword_results:
                        keyword_results[keyword] = []
                    keyword_results[keyword].append(name)
                else:
                    # 未匹配任何关键词，保存原始keyword和name
                    original_keyword = inst.get('matched_keyword', '').strip()
                    unmatched_results.append({
                        'keyword': original_keyword if original_keyword else '未知',
                        'name': name
                    })
    
    return keyword_results, unmatched_results


def load_url_scan_log(log_file: Path) -> Dict[str, Dict[str, Any]]:
    """
    加载方式2的 url_scan log
    
    Returns:
        {journal_name: {'results': [结果列表], 'url1': str, 'url2': str}}
    """
    if not log_file.exists():
        return None
    
    results = {}
    
    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['status'] == 'success' and row.get('results_json'):
                    journal_name = row['journal_name']
                    try:
                        results_list = json.loads(row['results_json'])
                        results[journal_name] = {
                            'results': results_list,
                            'url1': row.get('url1', ''),
                            'url2': row.get('url2', '')
                        }
                    except json.JSONDecodeError:
                        pass
    except Exception as e:
        print(f"[ERROR] Failed to load url_scan log: {e}", file=sys.stderr)
        return None
    
    return results


def load_search_log(log_file: Path) -> Dict[str, List[Dict[str, Any]]]:
    """
    加载方式3的 search log
    
    Returns:
        {journal_name: [结果列表]}
    """
    if not log_file.exists():
        return None
    
    results = {}
    
    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['status'] == 'success' and row.get('results_json'):
                    journal_name = row['journal_name']
                    try:
                        results_list = json.loads(row['results_json'])
                        results[journal_name] = results_list
                    except json.JSONDecodeError:
                        pass
    except Exception as e:
        print(f"[ERROR] Failed to load search log: {e}", file=sys.stderr)
        return None
    
    return results


# ========== 数据整合 ==========

def merge_list_values(values: List[Any]) -> str:
    """将列表值用`;`合并为字符串，支持字符串和列表"""
    if not values:
        return ""
    # 去重并保持顺序
    seen = set()
    unique_values = []
    for v in values:
        if not v:
            continue
        # 如果 v 是列表，先展开
        if isinstance(v, list):
            for item in v:
                if item and str(item).strip():
                    item_str = str(item).strip()
                    if item_str not in seen:
                        seen.add(item_str)
                        unique_values.append(item_str)
        else:
            v_str = str(v).strip()
            if v_str and v_str not in seen:
                seen.add(v_str)
                unique_values.append(v_str)
    return ";".join(unique_values)


def combine_data(
    df: pd.DataFrame,
    header_names: List[str],
    snapshot_dir: Path,
    url_scan_results: Dict[str, Dict[str, Any]],
    search_results: Dict[str, List[Dict[str, Any]]]
) -> List[Dict[str, Any]]:
    """
    整合所有数据
    
    Returns:
        输出行列表
    """
    output_rows = []
    
    # 找到关键列的索引
    try:
        journal_name_col = header_names.index('期刊名称')
        intro_url_col = header_names.index('期刊官方简介链接')
        host_url_col = header_names.index('主办单位官方链接')
        issn_col = header_names.index('ISSN')
        eissn_col = header_names.index('eISSN')
        manual_unit_col = header_names.index('人工判断单位')
        manual_sentence_col = header_names.index('人工判断关键语句')
    except ValueError as e:
        print(f"[ERROR] Required column not found in header: {e}", file=sys.stderr)
        sys.exit(1)
    
    # 处理每一行
    for idx, row in df.iterrows():
        journal_name = str(row.iloc[journal_name_col]).strip() if pd.notna(row.iloc[journal_name_col]) else ""
        intro_url = str(row.iloc[intro_url_col]).strip() if pd.notna(row.iloc[intro_url_col]) else ""
        host_url = str(row.iloc[host_url_col]).strip() if pd.notna(row.iloc[host_url_col]) else ""
        issn = str(row.iloc[issn_col]).strip() if pd.notna(row.iloc[issn_col]) else ""
        eissn = str(row.iloc[eissn_col]).strip() if pd.notna(row.iloc[eissn_col]) else ""
        manual_unit = str(row.iloc[manual_unit_col]).strip() if pd.notna(row.iloc[manual_unit_col]) else ""
        manual_sentence = str(row.iloc[manual_sentence_col]).strip() if pd.notna(row.iloc[manual_sentence_col]) else ""
        
        # 初始化输出行
        output_row = {
            # 原始数据
            '原始_期刊名称': journal_name,
            '原始_期刊官方简介链接': intro_url,
            '原始_主办单位官方链接': host_url,
            '原始_ISSN': issn,
            '原始_eISSN': eissn,
            
            # 人工查找
            '人工_人工判断单位': manual_unit,
            '人工_人工判断关键语句': manual_sentence,
            
            # AI精准操作提示（方式2）
            'AI精准_期刊名称': "",
            'AI精准_关联单位': "",
            'AI精准_关键句子': "",
            'AI精准_信息位置': "",
            'AI精准_来源链接1': "",
            'AI精准_来源链接2': "",
            
            # AI核心目标提示（方式3）
            'AI核心_期刊名称': "",
            'AI核心_主办单位': "",
            'AI核心_关键句子': "",
            'AI核心_判断依据': "",
            'AI核心_来源链接': "",
        }
        
        # 添加关键词列（方式1）
        for keyword in KEYWORD_COLUMNS:
            output_row[f'关键词_{keyword}'] = ""
        output_row['关键词_其他'] = ""
        
        # === 方式1：按关键词规则提取 ===
        urls = [intro_url, host_url]
        keyword_results, unmatched_results = load_extract_results(snapshot_dir, urls)
        
        for keyword_display in KEYWORD_COLUMNS:
            keyword_lower = keyword_display.lower()
            if keyword_lower in keyword_results:
                institutions = keyword_results[keyword_lower]
                output_row[f'关键词_{keyword_display}'] = merge_list_values(institutions)
        
        # 处理未匹配的项，格式为 "keyword: matches;..."
        if unmatched_results:
            other_items = []
            for item in unmatched_results:
                keyword = item['keyword']
                name = item['name']
                other_items.append(f"{keyword}: {name}")
            output_row['关键词_其他'] = ";".join(other_items)
        
        # === 方式2：AI精准操作提示 ===
        if url_scan_results and journal_name in url_scan_results:
            scan_data = url_scan_results[journal_name]
            results = scan_data['results']
            if results:
                # 合并多个结果
                output_row['AI精准_期刊名称'] = journal_name
                output_row['AI精准_关联单位'] = merge_list_values([r.get('关联单位', '') for r in results])
                output_row['AI精准_关键句子'] = merge_list_values([r.get('关键句子', '') for r in results])
                output_row['AI精准_信息位置'] = merge_list_values([r.get('信息位置', '') for r in results])
                
                # 来源链接（从 log 中读取）
                output_row['AI精准_来源链接1'] = scan_data['url1']
                output_row['AI精准_来源链接2'] = scan_data['url2']
        
        # === 方式3：AI核心目标提示 ===
        if search_results and journal_name in search_results:
            results = search_results[journal_name]
            if results:
                # 合并多个结果
                output_row['AI核心_期刊名称'] = journal_name
                output_row['AI核心_主办单位'] = merge_list_values([r.get('主办单位', '') for r in results])
                output_row['AI核心_关键句子'] = merge_list_values([r.get('关键句子', '') for r in results])
                output_row['AI核心_判断依据'] = merge_list_values([r.get('判断依据', '') for r in results])
                output_row['AI核心_来源链接'] = merge_list_values([r.get('来源链接', '') for r in results])
        
        output_rows.append(output_row)
    
    return output_rows


# ========== Excel 输出 ==========

def write_output_excel(output_rows: List[Dict[str, Any]], output_file: Path):
    """
    写入 Excel，带双层表头和颜色标记
    """
    try:
        # 创建 DataFrame
        df = pd.DataFrame(output_rows)
        
        # 确保输出目录存在
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # 写入 Excel（先不带表头）
        df.to_excel(output_file, index=False, header=False, startrow=2, engine='openpyxl')
        
        # 使用 openpyxl 添加双层表头
        wb = load_workbook(output_file)
        ws = wb.active
        ws.title = '整合输出'
        
        # 定义分组和字段
        groups = [
            {
                'name': '原始数据',
                'fields': ['期刊名称', '期刊官方简介链接', '主办单位官方链接', 'ISSN', 'eISSN'],
                'color': 'FACDED'
            },
            {
                'name': '人工查找',
                'fields': ['人工判断单位', '人工判断关键语句'],
                'color': '7CDED7'
            },
            {
                'name': 'AI精准操作提示',
                'fields': ['期刊名称', '关联单位', '关键句子', '信息位置', '来源链接1', '来源链接2'],
                'color': 'FFF258'
            },
            {
                'name': 'AI核心目标提示',
                'fields': ['期刊名称', '主办单位', '关键句子', '判断依据', '来源链接'],
                'color': 'BACEFD'
            },
            {
                'name': '按关键词规则提取',
                'fields': KEYWORD_COLUMNS + ['其他'],
                'color': '249087',
                'font_color': 'FFFFFF'
            }
        ]
        
        # 定义边框样式
        border_style = Border(
            left=Side(style='thin', color='666666'),
            right=Side(style='thin', color='666666'),
            top=Side(style='thin', color='666666'),
            bottom=Side(style='thin', color='666666')
        )
        
        # 定义对齐样式（居中）
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入第一行（分组）和第二行（字段）
        col_idx = 1
        for group in groups:
            start_col = col_idx
            end_col = col_idx + len(group['fields']) - 1
            
            # 第一行：分组名称（合并单元格）
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            cell = ws.cell(row=1, column=start_col, value=group['name'])
            cell.fill = PatternFill(start_color=group['color'], end_color=group['color'], fill_type='solid')
            cell.border = border_style
            cell.alignment = alignment_center
            
            # 设置字体颜色（如果需要）
            if group.get('font_color'):
                cell.font = Font(color=group['font_color'], bold=True)
            else:
                cell.font = Font(bold=True)
            
            # 第二行：字段名称
            for i, field in enumerate(group['fields']):
                cell = ws.cell(row=2, column=start_col + i, value=field)
                cell.fill = PatternFill(start_color=group['color'], end_color=group['color'], fill_type='solid')
                cell.border = border_style
                cell.alignment = alignment_center
                
                # 设置字体颜色
                if group.get('font_color'):
                    cell.font = Font(color=group['font_color'], bold=True)
                else:
                    cell.font = Font(bold=True)
            
            col_idx = end_col + 1
        
        # 冻结前两行
        ws.freeze_panes = 'A3'
        
        # 调整列宽
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            # 根据列内容调整宽度
            if col_idx <= 7:  # 原始数据 + 人工查找
                ws.column_dimensions[col_letter].width = 30
            elif col_idx <= 13:  # AI精准操作提示
                ws.column_dimensions[col_letter].width = 35
            elif col_idx <= 18:  # AI核心目标提示
                ws.column_dimensions[col_letter].width = 35
            else:  # 关键词列
                ws.column_dimensions[col_letter].width = 25
        
        # 保存
        wb.save(output_file)
        
        print(f"[OK] 输出文件已保存: {output_file}")
        print(f"     总行数: {len(output_rows)}")
        
    except Exception as e:
        print(f"[ERROR] Failed to write output Excel: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


# ========== 主函数 ==========

def main():
    parser = argparse.ArgumentParser(
        description="三种方法对比整合工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python combine_output3.py \\
    --input-excel journals.xlsx \\
    --sheet-name 0 \\
    --rows 2-99

  python combine_output3.py \\
    --input-excel journals.xlsx \\
    --sheet-name 0 \\
    --rows 2+

前置要求：
  三种方法必须已执行并生成对应的 log 文件：
  1. {excel_dir}/{excel_stem}-snapshot/extract-log.csv
  2. {excel_dir}/{excel_name}-url-scan-log.csv
  3. {excel_dir}/{excel_name}-search-log.csv
        """
    )
    
    parser.add_argument(
        '--input-excel',
        required=True,
        help='Excel 文件路径'
    )
    parser.add_argument(
        '--sheet-name',
        default=0,
        help='Sheet 名称或索引（默认 0，即第一个 sheet）'
    )
    parser.add_argument(
        '--rows',
        required=True,
        help='行范围，如 "2-99" 或 "2+"（从第2行开始到空行结束，第1行是表头）'
    )
    
    args = parser.parse_args()
    
    # 打印配置参数
    print("=" * 60)
    print("[CONFIG] 三种方法对比整合工具 - 启动参数")
    print("=" * 60)
    print(f"Excel 文件:    {args.input_excel}")
    print(f"Sheet 名称:    {args.sheet_name}")
    print(f"行范围:        {args.rows}")
    print("=" * 60)
    print()
    
    # 检查 Excel 文件
    excel_path = Path(args.input_excel)
    if not excel_path.exists():
        print(f"[ERROR] Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)
    
    # 解析参数
    try:
        # 处理 sheet_name（可能是数字或字符串）
        sheet_name = args.sheet_name
        try:
            sheet_name = int(sheet_name)
        except (ValueError, TypeError):
            pass
        
        # 解析行范围
        start_row, end_row = parse_rows_range(args.rows)
        
    except Exception as e:
        print(f"[ERROR] Invalid arguments: {e}", file=sys.stderr)
        sys.exit(1)
    
    # 确定三种方法的 log 文件路径
    snapshot_dir = excel_path.parent / f"{excel_path.stem}-snapshot"
    extract_log = snapshot_dir / "extract-log.csv"
    url_scan_log = excel_path.parent / f"{excel_path.name}-url-scan-log.csv"
    search_log = excel_path.parent / f"{excel_path.name}-search-log.csv"
    
    print(f"[CHECK] 检查 log 文件...")
    print(f"  方式1 (extract): {extract_log}")
    print(f"  方式2 (url_scan): {url_scan_log}")
    print(f"  方式3 (search): {search_log}")
    print()
    
    # 检查 log 文件是否存在
    missing_logs = []
    
    if not snapshot_dir.exists():
        missing_logs.append(f"  - snapshot 目录不存在: {snapshot_dir}")
    elif not extract_log.exists():
        missing_logs.append(f"  - extract-log.csv 不存在: {extract_log}")
    
    if not url_scan_log.exists():
        missing_logs.append(f"  - url-scan-log.csv 不存在: {url_scan_log}")
    
    if not search_log.exists():
        missing_logs.append(f"  - search-log.csv 不存在: {search_log}")
    
    if missing_logs:
        print("[ERROR] 缺少必需的 log 文件:", file=sys.stderr)
        for msg in missing_logs:
            print(msg, file=sys.stderr)
        print("\n请先运行对应的批处理脚本生成 log 文件。", file=sys.stderr)
        sys.exit(1)
    
    print("[OK] 所有 log 文件已就绪")
    print()
    
    # 读取输入 Excel
    print(f"[COMBINE] 读取 Excel 数据...")
    header_names, df = read_input_excel(excel_path, sheet_name, start_row, end_row)
    print(f"[COMBINE] 读取到 {len(df)} 行数据")
    print()
    
    # 加载三种方法的结果
    print(f"[COMBINE] 加载方式2结果 (url_scan)...")
    url_scan_results = load_url_scan_log(url_scan_log)
    if url_scan_results is None:
        print("[ERROR] 加载 url_scan log 失败", file=sys.stderr)
        sys.exit(1)
    print(f"[COMBINE] url_scan 结果: {len(url_scan_results)} 个期刊")
    
    print(f"[COMBINE] 加载方式3结果 (search)...")
    search_results = load_search_log(search_log)
    if search_results is None:
        print("[ERROR] 加载 search log 失败", file=sys.stderr)
        sys.exit(1)
    print(f"[COMBINE] search 结果: {len(search_results)} 个期刊")
    print()
    
    # 整合数据
    print(f"[COMBINE] 整合数据...")
    output_rows = combine_data(df, header_names, snapshot_dir, url_scan_results, search_results)
    print(f"[COMBINE] 生成 {len(output_rows)} 行输出")
    print()
    
    # 生成输出文件名（包含数据行数）
    timestamp = datetime.now().strftime("%y%m%d.%H%M%S")
    output_filename = f"{excel_path.stem}-{len(output_rows)}-{timestamp}.xlsx"
    output_file = excel_path.parent / output_filename
    
    # 写入 Excel
    print(f"[COMBINE] 写入输出文件...")
    write_output_excel(output_rows, output_file)
    
    # 统计信息
    print(f"\n[OK] 整合完成")
    print(f"     输出文件: {output_file}")


if __name__ == "__main__":
    main()

