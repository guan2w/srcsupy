#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
assemble.py

整合输入 Excel 文件和搜索结果日志，生成包含搜索结果的新 Excel 文件。

依赖：pip install openpyxl

使用示例：
    python assemble.py \
        --input-file=/path/to/file.xlsx \
        --sheet-name=Sheet1 \
        --header-row=2 \
        --rows=3+ \
        --top-n=3 \
        --columns=来源链接:url,域名:domain
"""

import argparse
import csv
import os
import re
import shutil
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# 兼容不同 Python 版本的 toml 库导入
try:
    import tomllib
except ImportError:
    try:
        import tomli as tomllib
    except ImportError:
        tomllib = None


# ==================== 工具函数 ====================

def log_print(*args, level="INFO", **kwargs):
    """通用日志打印函数"""
    import datetime as dt
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] [{level}]", *args, **kwargs)


def load_config(config_path: str) -> Dict[str, Any]:
    """读取配置文件"""
    cfg = {
        "exclude_url_pattern": None
    }
    if not os.path.exists(config_path) or tomllib is None:
        return cfg
    try:
        with open(config_path, "rb") as f:
            data = tomllib.load(f)
            assemble_cfg = data.get("assemble", {})
            cfg.update({k: v for k, v in assemble_cfg.items() if k in cfg})
    except Exception as e:
        log_print(f"解析 config.toml 出错：{e}", level="WARN")
    return cfg


def parse_rows_spec(spec: str, max_row: int, start_row: int) -> Tuple[int, int]:
    """解析行范围参数"""
    spec = spec.strip()
    if spec.endswith("+"):
        row_start = int(spec[:-1])
        row_end = max_row
    elif "-" in spec:
        parts = spec.split("-", 1)
        row_start, row_end = int(parts[0]), int(parts[1])
    else:
        row_start = int(spec)
        row_end = max_row
    
    if row_start < start_row:
        raise ValueError(f"rows 起始行 {row_start} 不能小于数据起始行 {start_row}")
    if row_end < row_start:
        raise ValueError(f"rows 范围非法：{row_start}-{row_end}")
    
    return row_start, min(row_end, max_row)


def parse_columns_spec(spec: str) -> List[Tuple[str, str]]:
    """
    解析 columns 参数
    格式: 来源链接:url,域名:domain
    返回: [(输出列名, JSON字段名), ...]
    """
    result = []
    for item in spec.split(","):
        item = item.strip()
        if not item:
            continue
        if ":" not in item:
            raise ValueError(f"columns 格式错误，缺少冒号: {item}")
        parts = item.split(":", 1)
        output_name = parts[0].strip()
        field_name = parts[1].strip()
        if not output_name or not field_name:
            raise ValueError(f"columns 格式错误: {item}")
        result.append((output_name, field_name))
    
    if not result:
        raise ValueError("columns 参数不能为空")
    
    return result


# ==================== Excel 处理 ====================

def read_meta_value(wb, key: str, meta_sheet: str = "meta") -> str:
    """
    从 meta sheet 读取指定 key 的 value
    meta 表结构：A1="key", B1="value"，数据从第2行开始
    """
    if meta_sheet not in wb.sheetnames:
        raise ValueError(f"未找到 meta sheet: {meta_sheet}")
    
    ws = wb[meta_sheet]
    
    # 遍历查找 key
    for row in range(2, ws.max_row + 1):
        cell_key = ws.cell(row=row, column=1).value
        if cell_key and str(cell_key).strip() == key:
            cell_value = ws.cell(row=row, column=2).value
            if not cell_value:
                raise ValueError(f"meta 表中 key='{key}' 的 value 为空，请填写")
            return str(cell_value).strip()
    
    raise ValueError(f"meta 表中未找到 key='{key}'，请添加")


def read_search_template(wb) -> str:
    """从 meta sheet 读取搜索模板（key='search'）"""
    return read_meta_value(wb, "search")


def extract_template_variables(template: str) -> List[str]:
    """从模板中提取变量名"""
    pattern = r'\{\{(\w+)\}\}'
    return re.findall(pattern, template)


def read_header_mapping(ws, header_row: int) -> Dict[str, int]:
    """读取表头，返回列名到列索引的映射（列索引从 0 开始，适配 iter_rows）"""
    header_map = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value:
            # 存储 0-based 索引，适配 iter_rows 返回的 tuple
            header_map[str(cell_value).strip()] = col - 1
    return header_map


# 预编译模板正则，用于快速渲染
_TEMPLATE_PATTERN = re.compile(r'\{\{(\w+)\}\}')


def render_template_fast(template: str, row_data: Dict[str, Any], pattern: re.Pattern = _TEMPLATE_PATTERN) -> str:
    """根据行数据渲染模板（优化版：正则一次替换）"""
    def replacer(match):
        key = match.group(1)
        value = row_data.get(key)
        return str(value).strip() if value is not None else ""
    return pattern.sub(replacer, template).strip()


def build_var_indices(header_map: Dict[str, int], variables: List[str]) -> List[Tuple[str, Optional[int]]]:
    """
    预计算变量到列索引的映射
    返回: [(变量名, 列索引或None), ...]
    """
    return [(var, header_map.get(var)) for var in variables]


def extract_row_data(row_tuple: tuple, var_indices: List[Tuple[str, Optional[int]]]) -> Dict[str, Any]:
    """从行元组中提取变量数据（优化版：使用预计算的索引）"""
    row_data = {}
    for var, col_idx in var_indices:
        if col_idx is not None and col_idx < len(row_tuple):
            row_data[var] = row_tuple[col_idx]
        else:
            row_data[var] = ""
    return row_data


# ==================== 日志处理 ====================

def load_search_log(log_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    加载搜索日志，返回按 query 索引的字典
    每个 query 对应一个结果列表，按 position 排序
    """
    log_data = defaultdict(list)
    
    if not os.path.exists(log_path):
        return log_data
    
    try:
        with open(log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                query = row.get("query", "").strip()
                if not query:
                    continue
                
                # 只收集有效结果（FOUND=true 且有 url）
                found = row.get("FOUND", "").lower() == "true"
                url = row.get("url", "").strip()
                
                if found and url:
                    try:
                        position = int(row.get("position", 0))
                    except (ValueError, TypeError):
                        position = 0
                    
                    log_data[query].append({
                        "position": position,
                        "url": url,
                        "displayed_url": row.get("displayed_url", ""),
                        "description": row.get("description", ""),
                        "title": row.get("title", ""),
                        "domain": row.get("domain", ""),
                        "number_of_results": row.get("number_of_results", ""),
                        "number_of_organic_results": row.get("number_of_organic_results", "")
                    })
    except Exception as e:
        log_print(f"读取搜索日志出错：{e}", level="ERROR")
    
    # 按 position 排序
    for query in log_data:
        log_data[query].sort(key=lambda x: x["position"])
    
    return log_data


# ==================== 主程序 ====================

def main():
    parser = argparse.ArgumentParser(description="整合搜索结果到 Excel 文件")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet-name", required=True, help="输入数据的 Sheet 名称")
    parser.add_argument("--header-row", type=int, default=1, help="表头行号（默认 1）")
    parser.add_argument("--rows", required=True, help="数据行范围，例如 3+ 或 3-100")
    parser.add_argument("--top-n", type=int, default=1, help="整合的搜索结果条数（默认 1）")
    parser.add_argument("--columns", required=True, help="输出列映射，如 来源链接:url,域名:domain")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    
    args = parser.parse_args()
    
    # 验证输入文件
    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path):
        log_print(f"输入文件不存在: {input_path}", level="ERROR")
        sys.exit(1)
    
    # 路径处理
    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    log_path = os.path.join(base_dir, f"{base_name}-search-log.csv")
    output_path = os.path.join(base_dir, f"{base_name}-assembled.xlsx")
    
    # 加载配置文件
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config.toml")
    if not os.path.exists(config_path):
        config_path = os.path.join(base_dir, "config.toml")
    cfg = load_config(config_path)
    
    # URL 排除模式
    exclude_pattern = cfg.get("exclude_url_pattern")
    exclude_regex = re.compile(exclude_pattern) if exclude_pattern else None
    
    # 解析 columns 参数
    try:
        columns_spec = parse_columns_spec(args.columns)
    except ValueError as e:
        log_print(f"columns 参数错误: {e}", level="ERROR")
        sys.exit(1)
    
    # 显示配置信息
    log_print("=" * 70)
    log_print("配置信息")
    log_print(f"  输入文件: {input_path}")
    log_print(f"  Sheet: {args.sheet_name}")
    log_print(f"  表头行: {args.header_row}")
    log_print(f"  数据行范围: {args.rows}")
    log_print(f"  Top-N: {args.top_n}")
    log_print(f"  输出列: {columns_spec}")
    log_print(f"  URL排除模式: {exclude_pattern or '无'}")
    log_print(f"  搜索日志: {log_path}")
    log_print(f"  输出文件: {output_path}")
    log_print("=" * 70)
    
    # 验证搜索日志
    if not os.path.exists(log_path):
        log_print(f"搜索日志文件不存在: {log_path}", level="ERROR")
        sys.exit(1)
    
    # 加载搜索日志（预索引，提升性能）
    log_print("正在加载搜索日志...")
    search_data = load_search_log(log_path)
    log_print(f"加载了 {len(search_data)} 条唯一查询的搜索结果")
    
    # 加载 Excel 文件（只读模式获取数据）
    log_print("正在加载 Excel 文件...")
    wb_read = load_workbook(input_path, read_only=True, data_only=True)
    
    # 读取模板
    template = read_search_template(wb_read)
    variables = extract_template_variables(template)
    log_print(f"搜索模板: {template}")
    log_print(f"模板变量: {variables}")
    
    # 验证 sheet
    if args.sheet_name not in wb_read.sheetnames:
        log_print(f"未找到 sheet: {args.sheet_name}", level="ERROR")
        sys.exit(1)
    
    ws_read = wb_read[args.sheet_name]
    
    # 读取表头映射
    header_map = read_header_mapping(ws_read, args.header_row)
    
    # 解析行范围
    data_start_row = args.header_row + 1
    start_row, end_row = parse_rows_spec(args.rows, ws_read.max_row, data_start_row)
    total_rows = end_row - start_row + 1
    log_print(f"处理行范围: {start_row}-{end_row} (共 {total_rows} 行)")
    
    # 关闭只读工作簿
    wb_read.close()
    
    # 复制输入文件到输出文件
    log_print("正在复制输入文件...")
    shutil.copy2(input_path, output_path)
    
    # 打开输出文件进行编辑
    wb_write = load_workbook(output_path)
    ws_write = wb_write[args.sheet_name]
    
    # 计算新列的起始位置
    original_max_col = ws_write.max_column
    
    # 生成新列标题（顺序 A：来源链接1, 域名1, 来源链接2, 域名2）
    new_headers = []
    for i in range(1, args.top_n + 1):
        for output_name, _ in columns_spec:
            new_headers.append(f"{output_name}{i}")
    
    # 写入新列标题
    for i, header in enumerate(new_headers):
        col_idx = original_max_col + 1 + i
        ws_write.cell(row=args.header_row, column=col_idx, value=header)
    
    log_print(f"新增列: {new_headers}")
    
    # 重新加载只读工作簿以读取数据
    wb_read = load_workbook(input_path, read_only=True, data_only=True)
    ws_read = wb_read[args.sheet_name]
    
    # 重新获取表头映射（0-based 索引）
    header_map = read_header_mapping(ws_read, args.header_row)
    
    # 预计算变量列索引（避免每行都查找）
    var_indices = build_var_indices(header_map, variables)
    
    # 预计算字段名列表（避免循环内重复解包）
    field_names = [field_name for _, field_name in columns_spec]
    num_fields = len(field_names)
    
    # 批量读取所有数据行到内存（核心优化）
    log_print("正在批量读取 Excel 数据...")
    all_rows = list(ws_read.iter_rows(
        min_row=start_row,
        max_row=end_row,
        values_only=True
    ))
    log_print(f"已读取 {len(all_rows)} 行数据到内存")
    
    # 处理每一行并收集写入数据
    matched_count = 0
    not_found_count = 0
    write_batch = []  # 收集所有写入操作: (row_idx, col_idx, value)
    
    col_offset = original_max_col + 1
    
    for idx, row_tuple in enumerate(all_rows):
        row_idx = start_row + idx
        
        # 从行元组中提取变量数据
        row_data = extract_row_data(row_tuple, var_indices)
        query = render_template_fast(template, row_data)
        
        if not query:
            if args.debug:
                log_print(f"行 {row_idx} 渲染后为空，跳过", level="DEBUG")
            not_found_count += 1
            continue
        
        # 查找搜索结果（O(1) 字典查找）
        results = search_data.get(query)
        
        if not results:
            if args.debug:
                log_print(f"行 {row_idx} 未找到搜索结果: {query[:40]}...", level="DEBUG")
            not_found_count += 1
            continue
        
        # 根据 exclude_url_pattern 过滤结果
        if exclude_regex:
            results = [r for r in results if not exclude_regex.search(r.get("url", ""))]
            if not results:
                if args.debug:
                    log_print(f"行 {row_idx} 结果被排除: {query[:40]}...", level="DEBUG")
                not_found_count += 1
                continue
        
        matched_count += 1
        
        # 收集写入数据（延迟写入，提升性能）
        for result_idx in range(min(args.top_n, len(results))):
            result = results[result_idx]
            base_col = col_offset + result_idx * num_fields
            for field_idx, field_name in enumerate(field_names):
                value = result.get(field_name, "")
                if value:  # 只收集非空值
                    write_batch.append((row_idx, base_col + field_idx, value))
        
        # 进度显示
        if (idx + 1) % 500 == 0:
            progress = (idx + 1) / total_rows * 100
            log_print(f"处理进度: {idx + 1}/{total_rows} ({progress:.1f}%)")
    
    # 批量写入 Excel
    log_print(f"正在批量写入 {len(write_batch)} 个单元格...")
    for row_idx, col_idx, value in write_batch:
        ws_write.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    for i, header in enumerate(new_headers):
        col_idx = original_max_col + 1 + i
        col_letter = ws_write.cell(row=1, column=col_idx).column_letter
        # 根据内容类型设置列宽
        if "url" in header.lower() or "链接" in header:
            ws_write.column_dimensions[col_letter].width = 60
        else:
            ws_write.column_dimensions[col_letter].width = 20
    
    # 保存
    log_print("正在保存输出文件...")
    wb_write.save(output_path)
    wb_write.close()
    wb_read.close()
    
    # 显示统计
    log_print("=" * 70)
    log_print("任务完成")
    log_print(f"  处理行数: {total_rows}")
    log_print(f"  匹配成功: {matched_count}")
    log_print(f"  未找到结果: {not_found_count}")
    log_print(f"  输出文件: {output_path}")
    log_print("=" * 70)


if __name__ == "__main__":
    main()

