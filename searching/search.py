#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
search.py

读取 Excel 输入文件，根据 template sheet 中的模板构造搜索关键词，
调用 ScrapingBee Google Search API 执行搜索，将结果记录到 CSV 文件。

依赖：pip install openpyxl requests

使用示例：
    python search.py \
        --input-file=/path/to/file.xlsx \
        --sheet-name=Sheet1 \
        --header-row=2 \
        --rows=3+ \
        --top-n=5 \
        --concurrency=10

配置文件：config.toml 或环境变量 SCRAPINGBEE_API_KEY
"""

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from typing import Any, Dict, List, Optional, Set, Tuple

import requests
from openpyxl import load_workbook

# 兼容不同 Python 版本的 toml 库导入
try:
    import tomllib
except ImportError:
    try:
        import tomli as tomllib
    except ImportError:
        tomllib = None

# 全局常量
LOG_HEADER = [
    "row_number", "query", "search_time", "duration_ms", "FOUND", "ERROR",
    "number_of_results", "number_of_organic_results",
    "position", "url", "displayed_url", "description", "title", "domain"
]

# ==================== 工具函数 ====================

def log_print(*args, level="INFO", **kwargs):
    """通用日志打印函数，带时间戳和级别标识"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] [{level}]", *args, **kwargs)


def load_env_file(path: str = ".env"):
    """加载 .env 文件到环境变量"""
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key, value = key.strip(), value.strip()
                if (value.startswith('"') and value.endswith('"')) or \
                   (value.startswith("'") and value.endswith("'")):
                    value = value[1:-1]
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception as e:
        log_print(f"加载 .env 文件出错：{e}", level="WARN")


def load_config(config_path: str) -> Dict[str, Any]:
    """读取配置，默认值兜底"""
    cfg = {
        "timeout_seconds": 120,
        "concurrency": 1,
        "retry_times": 1,
        "proxy": None
    }
    if not os.path.exists(config_path) or tomllib is None:
        return cfg
    try:
        with open(config_path, "rb") as f:
            data = tomllib.load(f)
            bee_cfg = data.get("scrapingbee", {})
            cfg.update({k: v for k, v in bee_cfg.items() if k in cfg})
    except Exception as e:
        log_print(f"解析 config.toml 出错，使用默认配置。错误：{e}", level="WARN")
    return cfg


def parse_rows_spec(spec: str, max_row: int, start_row: int) -> Tuple[int, int]:
    """解析行范围参数"""
    spec = spec.strip()
    if spec.endswith("+"):
        row_start = int(spec[:-1])
        row_end = max_row
    elif "-" in spec:
        parts = spec.split("-", 1)
        row_start, row_end = int(parts[0]), int(parts[1])
    else:
        row_start = int(spec)
        row_end = max_row
    
    if row_start < start_row:
        raise ValueError(f"rows 起始行 {row_start} 不能小于数据起始行 {start_row}")
    if row_end < row_start:
        raise ValueError(f"rows 范围非法：{row_start}-{row_end}")
    
    return row_start, min(row_end, max_row)


# ==================== Excel 处理 ====================

def read_template(wb, template_sheet: str = "template") -> str:
    """从 template sheet 读取搜索模板"""
    if template_sheet not in wb.sheetnames:
        raise ValueError(f"未找到 template sheet: {template_sheet}")
    
    ws = wb[template_sheet]
    template = ws.cell(row=2, column=1).value
    
    if not template:
        raise ValueError("template sheet A2 单元格为空，请填写搜索模板")
    
    return str(template).strip()


def extract_template_variables(template: str) -> List[str]:
    """从模板中提取变量名"""
    # 匹配 {{variable}} 格式
    pattern = r'\{\{(\w+)\}\}'
    return re.findall(pattern, template)


def read_header_mapping(ws, header_row: int) -> Dict[str, int]:
    """读取表头，返回列名到列索引的映射"""
    header_map = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value:
            header_map[str(cell_value).strip()] = col
    return header_map


def render_template(template: str, row_data: Dict[str, Any]) -> str:
    """根据行数据渲染模板"""
    result = template
    for key, value in row_data.items():
        placeholder = "{{" + key + "}}"
        if placeholder in result:
            # 处理 None 和空值
            str_value = str(value).strip() if value is not None else ""
            result = result.replace(placeholder, str_value)
    return result.strip()


def read_row_data(ws, row_idx: int, header_map: Dict[str, int], variables: List[str]) -> Dict[str, Any]:
    """读取指定行的数据"""
    row_data = {}
    for var in variables:
        if var in header_map:
            col_idx = header_map[var]
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[var] = cell_value
        else:
            row_data[var] = ""
    return row_data


# ==================== ScrapingBee API 调用 ====================

def search_google(api_key: str, query: str, timeout: int, proxies: Optional[Dict] = None) -> Tuple[Optional[Dict], Optional[str], float]:
    """
    执行 Google 搜索
    返回: (搜索结果字典, 错误信息, 耗时秒数)
    """
    url = "https://app.scrapingbee.com/api/v1/google"
    params = {
        "api_key": api_key,
        "search": query,
        "language": "zh-cn",
        "country_code": "cn"
    }
    
    start = time.monotonic()
    try:
        response = requests.get(url, params=params, timeout=timeout, proxies=proxies)
        duration = time.monotonic() - start
        
        if response.status_code != 200:
            return None, f"HTTP {response.status_code}: {response.text[:500]}", duration
        
        data = response.json()
        return data, None, duration
        
    except requests.exceptions.Timeout:
        return None, "请求超时", time.monotonic() - start
    except requests.exceptions.RequestException as e:
        return None, f"请求异常: {e}", time.monotonic() - start
    except json.JSONDecodeError as e:
        return None, f"JSON 解析错误: {e}", time.monotonic() - start


def retry_search(api_key: str, query: str, timeout: int, retry_times: int, proxies: Optional[Dict] = None) -> Tuple[Optional[Dict], Optional[str], float]:
    """带重试的搜索"""
    last_error = None
    total_duration = 0
    
    for attempt in range(1, retry_times + 2):  # +2 因为第一次不算重试
        result, error, duration = search_google(api_key, query, timeout, proxies)
        total_duration += duration
        
        if error is None:
            return result, None, total_duration
        
        last_error = error
        if attempt <= retry_times:
            log_print(f"搜索失败，尝试 {attempt}/{retry_times + 1}，错误: {error}", level="WARN")
            time.sleep(min(1.0 * attempt, 5.0))
    
    return None, last_error, total_duration


# ==================== 日志处理 ====================

def ensure_log_header(log_path: str, header: List[str]):
    """创建日志文件表头"""
    if not os.path.exists(log_path):
        with open(log_path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(header)


def load_existing_queries(log_path: str) -> Set[str]:
    """加载已搜索的查询词集合"""
    existing = set()
    if not os.path.exists(log_path):
        return existing
    
    try:
        with open(log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                query = row.get("query", "").strip()
                if query:
                    existing.add(query)
    except Exception as e:
        log_print(f"读取现有日志出错：{e}", level="WARN")
    
    return existing


def write_search_results(
    writer,
    row_number: int,
    query: str,
    search_time: str,
    duration_ms: int,
    result: Optional[Dict],
    error: Optional[str],
    top_n: int
):
    """
    将搜索结果写入日志
    每条 organic_result 写一行
    """
    if error:
        # 搜索出错，写一行
        writer.writerow([
            row_number, query, search_time, duration_ms,
            "false", error,
            "", "",  # number_of_results, number_of_organic_results
            "", "", "", "", "", ""  # position, url, displayed_url, description, title, domain
        ])
        return
    
    # 提取元数据
    meta = result.get("meta_data", {})
    number_of_results = meta.get("number_of_results", 0)
    number_of_organic_results = meta.get("number_of_organic_results", 0)
    
    organic_results = result.get("organic_results", [])
    
    if not organic_results:
        # 无搜索结果，写一行
        writer.writerow([
            row_number, query, search_time, duration_ms,
            "false", "",
            number_of_results, number_of_organic_results,
            "", "", "", "", "", ""
        ])
        return
    
    # 根据 top_n 限制结果数量
    if top_n > 0:
        organic_results = organic_results[:top_n]
    
    # 每条结果写一行
    for item in organic_results:
        writer.writerow([
            row_number,
            query,
            search_time,
            duration_ms,
            "true",
            "",
            number_of_results,
            number_of_organic_results,
            item.get("position", ""),
            item.get("url", ""),
            item.get("displayed_url", ""),
            item.get("description", ""),
            item.get("title", ""),
            item.get("domain", "")
        ])


# ==================== 主程序 ====================

def main():
    parser = argparse.ArgumentParser(description="Excel 驱动的 ScrapingBee 搜索脚本")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet-name", required=True, help="输入数据的 Sheet 名称")
    parser.add_argument("--header-row", type=int, default=1, help="表头行号（默认 1）")
    parser.add_argument("--rows", required=True, help="数据行范围，例如 3+ 或 3-100")
    parser.add_argument("--top-n", type=int, default=0, help="保留前 N 条结果（0 表示全部）")
    parser.add_argument("--concurrency", type=int, default=0, help="并发数（0 使用配置文件值）")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    
    args = parser.parse_args()
    
    # 验证输入文件
    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path):
        log_print(f"输入文件不存在: {input_path}", level="ERROR")
        sys.exit(1)
    
    # 路径处理
    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    log_path = os.path.join(base_dir, f"{base_name}-search-log.csv")
    
    # 配置文件查找
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config.toml")
    if not os.path.exists(config_path):
        config_path = os.path.join(base_dir, "config.toml")
    
    # 加载环境变量
    load_env_file(os.path.join(script_dir, ".env"))
    load_env_file(os.path.join(base_dir, ".env"))
    load_env_file()
    
    # 加载配置
    cfg = load_config(config_path)
    
    # 并发数优先使用命令行参数
    concurrency = args.concurrency if args.concurrency > 0 else cfg["concurrency"]
    concurrency = max(1, concurrency)
    
    # 代理设置
    proxies = None
    if cfg.get("proxy"):
        proxies = {"http": cfg["proxy"], "https": cfg["proxy"]}
    
    # API 密钥
    api_key = os.environ.get("SCRAPINGBEE_API_KEY", "").strip()
    if not api_key:
        log_print("未找到 SCRAPINGBEE_API_KEY 环境变量", level="ERROR")
        sys.exit(1)
    
    # 显示配置信息
    log_print("=" * 70)
    log_print("配置信息")
    log_print(f"  输入文件: {input_path}")
    log_print(f"  Sheet: {args.sheet_name}")
    log_print(f"  表头行: {args.header_row}")
    log_print(f"  数据行范围: {args.rows}")
    log_print(f"  Top-N: {args.top_n if args.top_n > 0 else '全部'}")
    log_print(f"  并发数: {concurrency}")
    log_print(f"  日志文件: {log_path}")
    log_print(f"  配置文件: {config_path}")
    log_print("=" * 70)
    
    # 加载 Excel 文件
    log_print("正在加载 Excel 文件...")
    wb = load_workbook(input_path, read_only=True, data_only=True)
    
    # 读取模板
    template = read_template(wb)
    variables = extract_template_variables(template)
    log_print(f"搜索模板: {template}")
    log_print(f"模板变量: {variables}")
    
    # 验证 sheet
    if args.sheet_name not in wb.sheetnames:
        log_print(f"未找到 sheet: {args.sheet_name}", level="ERROR")
        sys.exit(1)
    
    ws = wb[args.sheet_name]
    
    # 读取表头映射
    header_map = read_header_mapping(ws, args.header_row)
    log_print(f"表头列映射: {header_map}")
    
    # 验证模板变量是否都在表头中
    missing_vars = [v for v in variables if v not in header_map]
    if missing_vars:
        log_print(f"模板变量在表头中未找到: {missing_vars}", level="ERROR")
        sys.exit(1)
    
    # 解析行范围
    data_start_row = args.header_row + 1
    start_row, end_row = parse_rows_spec(args.rows, ws.max_row, data_start_row)
    log_print(f"处理行范围: {start_row}-{end_row} (共 {end_row - start_row + 1} 行)")
    
    # 初始化日志文件
    ensure_log_header(log_path, LOG_HEADER)
    
    # 加载已搜索的查询词
    existing_queries = load_existing_queries(log_path)
    log_print(f"已存在的查询记录: {len(existing_queries)} 条")
    
    # 构建待处理任务
    tasks = []
    for row_idx in range(start_row, end_row + 1):
        row_data = read_row_data(ws, row_idx, header_map, variables)
        query = render_template(template, row_data)
        
        if not query.strip():
            if args.debug:
                log_print(f"行 {row_idx} 渲染后为空，跳过", level="DEBUG")
            continue
        
        if query in existing_queries:
            if args.debug:
                log_print(f"行 {row_idx} 查询已存在，跳过: {query[:50]}...", level="DEBUG")
            continue
        
        tasks.append((row_idx, query))
    
    if not tasks:
        log_print("所有行已处理完成，无需搜索")
        wb.close()
        return
    
    log_print(f"待处理任务数: {len(tasks)}")
    
    # 线程安全的日志写入
    log_lock = Lock()
    log_file = open(log_path, "a", newline="", encoding="utf-8")
    log_writer = csv.writer(log_file)
    
    # 进度统计
    processed = 0
    success_count = 0
    error_count = 0
    stats_lock = Lock()
    
    def process_task(task: Tuple[int, str]):
        nonlocal processed, success_count, error_count
        
        row_idx, query = task
        
        # 执行搜索
        result, error, duration = retry_search(
            api_key, query, cfg["timeout_seconds"], cfg["retry_times"], proxies
        )
        
        search_time = dt.datetime.now().astimezone().isoformat()
        duration_ms = int(duration * 1000)
        
        # 写入日志（线程安全）
        with log_lock:
            write_search_results(
                log_writer, row_idx, query, search_time, duration_ms,
                result, error, args.top_n
            )
            log_file.flush()
        
        # 更新统计
        with stats_lock:
            processed += 1
            if error:
                error_count += 1
            else:
                success_count += 1
            
            # 进度显示
            percentage = processed / len(tasks) * 100
            status = "✗" if error else "✓"
            log_print(f"[{processed}/{len(tasks)}] ({percentage:.1f}%) {status} 行 {row_idx}: {query[:40]}...")
    
    # 并发执行
    log_print(f"开始搜索，并发数: {concurrency}")
    
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = [executor.submit(process_task, task) for task in tasks]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                log_print(f"任务执行异常: {e}", level="ERROR")
    
    # 清理
    log_file.close()
    wb.close()
    
    # 显示统计
    log_print("=" * 70)
    log_print("任务完成")
    log_print(f"  总任务: {len(tasks)}")
    log_print(f"  成功: {success_count}")
    log_print(f"  失败: {error_count}")
    log_print(f"  日志文件: {log_path}")
    log_print("=" * 70)


if __name__ == "__main__":
    main()

