#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract-assemble.py

整合输入 Excel 文件和提取结果日志，在每个 URL 列右侧插入提取结果列。

依赖：pip install openpyxl

使用示例：
    python extract-assemble.py \
        --input-file=/path/to/file.xlsx \
        --sheet-name=Sheet1 \
        --header-row=2 \
        --rows=3+ \
        --url-columns=来源链接1,来源链接2
"""

import argparse
import csv
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from copy import copy
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ==================== 常量定义 ====================

# 列插入模式
INSERT_MODE_AFTER_URL = "after_url"  # 在每个 URL 列右侧插入
INSERT_MODE_APPEND = "append"        # 在所有列末尾追加

# 当前使用的插入模式（可通过配置修改）
CURRENT_INSERT_MODE = INSERT_MODE_AFTER_URL


# ==================== 工具函数 ====================

# Excel 非法字符正则表达式（控制字符，除了 tab \x09、换行 \x0a、回车 \x0d）
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def clean_illegal_characters(value: Any) -> Any:
    """清理 Excel 不允许的非法字符"""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def log_print(*args, level="INFO", **kwargs):
    """通用日志打印函数"""
    import datetime as dt
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] [{level}]", *args, **kwargs)


def parse_rows_spec(spec: str, max_row: int, start_row: int) -> Tuple[int, int]:
    """解析行范围参数"""
    spec = spec.strip()
    if spec.endswith("+"):
        row_start = int(spec[:-1])
        row_end = max_row
    elif "-" in spec:
        parts = spec.split("-", 1)
        row_start, row_end = int(parts[0]), int(parts[1])
    else:
        row_start = int(spec)
        row_end = max_row
    
    if row_start < start_row:
        raise ValueError(f"rows 起始行 {row_start} 不能小于数据起始行 {start_row}")
    if row_end < row_start:
        raise ValueError(f"rows 范围非法：{row_start}-{row_end}")
    
    return row_start, min(row_end, max_row)


def parse_url_columns(spec: str) -> List[str]:
    """解析 URL 列参数"""
    result = []
    for item in spec.split(","):
        item = item.strip()
        if item:
            result.append(item)
    if not result:
        raise ValueError("url-columns 参数不能为空")
    return result


# ==================== Excel 处理 ====================

def read_meta_value(wb, key: str, meta_sheet: str = "meta") -> str:
    """
    从 meta sheet 读取指定 key 的 value
    meta 表结构：A1="key", B1="value"，数据从第2行开始
    """
    if meta_sheet not in wb.sheetnames:
        raise ValueError(f"未找到 meta sheet: {meta_sheet}")
    
    ws = wb[meta_sheet]
    
    # 遍历查找 key
    for row in range(2, ws.max_row + 1):
        cell_key = ws.cell(row=row, column=1).value
        if cell_key and str(cell_key).strip() == key:
            cell_value = ws.cell(row=row, column=2).value
            if not cell_value:
                raise ValueError(f"meta 表中 key='{key}' 的 value 为空，请填写")
            return str(cell_value).strip()
    
    raise ValueError(f"meta 表中未找到 key='{key}'，请添加")


def read_extract_rules(wb) -> Dict[str, str]:
    """从 meta sheet 读取提取规则（key='ai_extract_rules'），获取字段列表"""
    rules_str = read_meta_value(wb, "ai_extract_rules")
    
    try:
        rules = json.loads(rules_str)
        if not isinstance(rules, dict):
            raise ValueError("提取规则必须是 JSON 对象")
        return rules
    except json.JSONDecodeError as e:
        raise ValueError(f"提取规则 JSON 解析失败: {e}")


def read_header_mapping(ws, header_row: int, use_iter: bool = False) -> Dict[str, int]:
    """
    读取表头，返回列名到列索引的映射
    use_iter=True 时使用 iter_rows（适用于 read_only 模式），返回 0-based 索引
    use_iter=False 时使用 cell（适用于写入模式），返回 1-based 索引
    """
    header_map = {}
    if use_iter:
        # read_only 模式：使用 iter_rows 批量读取，0-based 索引
        for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
            for col_idx, cell_value in enumerate(row):
                if cell_value:
                    header_map[str(cell_value).strip()] = col_idx  # 0-based
    else:
        # 写入模式：使用 cell，1-based 索引
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                header_map[str(cell_value).strip()] = col  # 1-based
    return header_map


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


# ==================== 日志处理 ====================

def load_extract_log(log_path: str) -> Dict[str, Dict[str, Any]]:
    """
    加载提取日志，返回按 URL 索引的字典
    每个 URL 对应其提取结果
    """
    log_data = {}
    
    if not os.path.exists(log_path):
        return log_data
    
    try:
        with open(log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            
            # 获取提取结果字段（排除基础字段）
            base_fields = {"url", "url_column", "row_number", "ai_extract_rules", "extract_time", "duration_ms", "SUCCESS", "ERROR"}
            result_fields = [f for f in fieldnames if f not in base_fields]
            
            for row in reader:
                url = row.get("url", "").strip()
                if not url:
                    continue
                
                success = row.get("SUCCESS", "").lower() == "true"
                
                # 只收集成功的结果
                if success:
                    result = {field: row.get(field, "") for field in result_fields}
                    log_data[url] = result
    except Exception as e:
        log_print(f"读取提取日志出错：{e}", level="ERROR")
    
    return log_data


def get_result_fields_from_log(log_path: str) -> List[str]:
    """从日志文件获取提取结果字段列表"""
    if not os.path.exists(log_path):
        return []
    
    try:
        with open(log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            
            base_fields = {"url", "url_column", "row_number", "ai_extract_rules", "extract_time", "duration_ms", "SUCCESS", "ERROR"}
            return [f for f in fieldnames if f not in base_fields]
    except Exception:
        return []


# ==================== 列插入策略 ====================

class ColumnInsertStrategy:
    """列插入策略基类"""
    
    def calculate_new_columns(
        self,
        url_columns: List[str],
        result_fields: List[str],
        header_map: Dict[str, int]
    ) -> List[Tuple[int, str, str]]:
        """
        计算需要插入的新列
        返回: [(插入位置, 列名, 对应的url_column), ...]
        """
        raise NotImplementedError


class InsertAfterUrlStrategy(ColumnInsertStrategy):
    """在每个 URL 列右侧插入策略"""
    
    def calculate_new_columns(
        self,
        url_columns: List[str],
        result_fields: List[str],
        header_map: Dict[str, int]
    ) -> List[Tuple[int, str, str]]:
        """
        在每个 URL 列右侧插入对应的结果列
        从右往左处理，避免列索引错乱
        """
        # 按列索引降序排列 URL 列（从右往左处理）
        url_cols_with_idx = [(col, header_map[col]) for col in url_columns]
        url_cols_with_idx.sort(key=lambda x: x[1], reverse=True)
        
        new_columns = []
        for url_col, col_idx in url_cols_with_idx:
            # 在该 URL 列右侧插入结果字段（倒序插入，保持字段顺序）
            for field in reversed(result_fields):
                new_col_name = f"{url_col}-{field}"
                insert_pos = col_idx + 1  # 在 URL 列右侧
                new_columns.append((insert_pos, new_col_name, url_col))
        
        return new_columns


class AppendStrategy(ColumnInsertStrategy):
    """在所有列末尾追加策略"""
    
    def calculate_new_columns(
        self,
        url_columns: List[str],
        result_fields: List[str],
        header_map: Dict[str, int]
    ) -> List[Tuple[int, str, str]]:
        """在所有原始列后追加"""
        max_col = max(header_map.values()) if header_map else 0
        new_columns = []
        
        insert_pos = max_col + 1
        for url_col in url_columns:
            for field in result_fields:
                new_col_name = f"{url_col}-{field}"
                new_columns.append((insert_pos, new_col_name, url_col))
                insert_pos += 1
        
        return new_columns


def get_insert_strategy(mode: str = CURRENT_INSERT_MODE) -> ColumnInsertStrategy:
    """获取列插入策略"""
    if mode == INSERT_MODE_AFTER_URL:
        return InsertAfterUrlStrategy()
    elif mode == INSERT_MODE_APPEND:
        return AppendStrategy()
    else:
        raise ValueError(f"未知的插入模式: {mode}")


# ==================== 主程序 ====================

def main():
    parser = argparse.ArgumentParser(description="整合提取结果到 Excel 文件")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet-name", required=True, help="输入数据的 Sheet 名称")
    parser.add_argument("--header-row", type=int, default=1, help="表头行号（默认 1）")
    parser.add_argument("--rows", required=True, help="数据行范围，例如 3+ 或 3-100")
    parser.add_argument("--url-columns", required=True, help="URL 列名，多个用逗号分隔")
    parser.add_argument("--insert-mode", default=CURRENT_INSERT_MODE, 
                        choices=[INSERT_MODE_AFTER_URL, INSERT_MODE_APPEND],
                        help=f"列插入模式（默认 {CURRENT_INSERT_MODE}）")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    
    args = parser.parse_args()
    
    # 验证输入文件
    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path):
        log_print(f"输入文件不存在: {input_path}", level="ERROR")
        sys.exit(1)
    
    # 路径处理
    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    log_path = os.path.join(base_dir, f"{base_name}-extract-log.csv")
    output_path = os.path.join(base_dir, f"{base_name}-extracted.xlsx")
    
    # 解析 URL 列
    try:
        url_columns = parse_url_columns(args.url_columns)
    except ValueError as e:
        log_print(f"url-columns 参数错误: {e}", level="ERROR")
        sys.exit(1)
    
    # 显示配置信息
    log_print("=" * 70)
    log_print("配置信息")
    log_print(f"  输入文件: {input_path}")
    log_print(f"  Sheet: {args.sheet_name}")
    log_print(f"  表头行: {args.header_row}")
    log_print(f"  数据行范围: {args.rows}")
    log_print(f"  URL 列: {url_columns}")
    log_print(f"  插入模式: {args.insert_mode}")
    log_print(f"  提取日志: {log_path}")
    log_print(f"  输出文件: {output_path}")
    log_print("=" * 70)
    
    # 验证提取日志
    if not os.path.exists(log_path):
        log_print(f"提取日志文件不存在: {log_path}", level="ERROR")
        sys.exit(1)
    
    # 加载提取日志（预索引）
    log_print("正在加载提取日志...")
    extract_data = load_extract_log(log_path)
    log_print(f"加载了 {len(extract_data)} 条成功的提取记录")
    
    # 获取结果字段列表
    result_fields = get_result_fields_from_log(log_path)
    if not result_fields:
        log_print("提取日志中未找到结果字段", level="ERROR")
        sys.exit(1)
    log_print(f"提取结果字段: {result_fields}")
    
    # 加载 Excel 文件（只读模式获取数据）
    log_print("正在加载 Excel 文件...")
    wb_read = load_workbook(input_path, read_only=True, data_only=True)
    
    # 验证 sheet
    if args.sheet_name not in wb_read.sheetnames:
        log_print(f"未找到 sheet: {args.sheet_name}", level="ERROR")
        sys.exit(1)
    
    ws_read = wb_read[args.sheet_name]
    
    # 读取表头映射（read_only 模式，使用 iter_rows，0-based 索引）
    header_map_read = read_header_mapping(ws_read, args.header_row, use_iter=True)
    
    # 验证 URL 列
    missing_cols = [col for col in url_columns if col not in header_map_read]
    if missing_cols:
        log_print(f"URL 列在表头中未找到: {missing_cols}", level="ERROR")
        sys.exit(1)
    
    # 解析行范围
    data_start_row = args.header_row + 1
    start_row, end_row = parse_rows_spec(args.rows, ws_read.max_row, data_start_row)
    total_rows = end_row - start_row + 1
    log_print(f"处理行范围: {start_row}-{end_row} (共 {total_rows} 行)")
    
    # 获取 URL 列的索引（0-based）
    url_col_indices = [(col, header_map_read[col]) for col in url_columns]
    
    # 批量读取所有数据行到内存（核心优化：避免逐行 cell 访问）
    log_print("正在批量读取 Excel 数据...")
    all_rows = list(ws_read.iter_rows(min_row=start_row, max_row=end_row, values_only=True))
    log_print(f"已读取 {len(all_rows)} 行数据到内存")
    
    wb_read.close()
    
    # 从批量读取的数据中提取 URL
    url_data = {}  # {row_idx: {url_column: url}}
    for idx, row_tuple in enumerate(all_rows):
        row_idx = start_row + idx
        url_data[row_idx] = {}
        for url_col, col_idx in url_col_indices:
            if col_idx < len(row_tuple):
                url = row_tuple[col_idx]
                if url:
                    url_data[row_idx][url_col] = str(url).strip()
    
    # 复制输入文件到输出文件
    log_print("正在复制输入文件...")
    shutil.copy2(input_path, output_path)
    
    # 打开输出文件进行编辑
    wb_write = load_workbook(output_path)
    ws_write = wb_write[args.sheet_name]
    
    # 重新获取表头映射（写入模式）
    header_map = read_header_mapping(ws_write, args.header_row)
    
    # 获取插入策略
    strategy = get_insert_strategy(args.insert_mode)
    
    # 计算需要插入的新列
    new_columns = strategy.calculate_new_columns(url_columns, result_fields, header_map)
    log_print(f"需要插入 {len(new_columns)} 个新列")
    
    if args.insert_mode == INSERT_MODE_AFTER_URL:
        # 方式 A：在 URL 列右侧插入（从右往左处理）
        # new_columns 已经按从右往左顺序排列
        
        # 记录新列的最终位置映射：{(url_col, field): final_col_idx}
        new_col_positions = {}
        
        for insert_pos, new_col_name, url_col in new_columns:
            # 插入空列
            ws_write.insert_cols(insert_pos)
            
            # 写入列标题
            ws_write.cell(row=args.header_row, column=insert_pos, value=new_col_name)
            
            # 提取字段名
            field = new_col_name.replace(f"{url_col}-", "")
            new_col_positions[(url_col, field)] = insert_pos
            
            if args.debug:
                log_print(f"插入列: {new_col_name} 在位置 {insert_pos}", level="DEBUG")
        
        # 更新表头映射（插入列后索引会变化）
        header_map = read_header_mapping(ws_write, args.header_row)
        
        # 填充数据
        matched_count = 0
        for row_idx in range(start_row, end_row + 1):
            row_urls = url_data.get(row_idx, {})
            
            for url_col in url_columns:
                url = row_urls.get(url_col)
                if not url:
                    continue
                
                result = extract_data.get(url)
                if not result:
                    continue
                
                matched_count += 1
                
                # 填充该 URL 对应的结果列
                for field in result_fields:
                    # 找到该列的位置
                    col_name = f"{url_col}-{field}"
                    if col_name in header_map:
                        col_idx = header_map[col_name]
                        value = result.get(field, "")
                        # 清理非法字符后写入
                        ws_write.cell(row=row_idx, column=col_idx, value=clean_illegal_characters(value))
            
            # 进度显示
            if (row_idx - start_row + 1) % 100 == 0:
                progress = (row_idx - start_row + 1) / total_rows * 100
                log_print(f"处理进度: {row_idx - start_row + 1}/{total_rows} ({progress:.1f}%)")
    
    else:
        # 方式 B：在末尾追加
        matched_count = 0
        
        for insert_pos, new_col_name, url_col in new_columns:
            # 写入列标题
            ws_write.cell(row=args.header_row, column=insert_pos, value=new_col_name)
        
        # 更新表头映射
        header_map = read_header_mapping(ws_write, args.header_row)
        
        # 填充数据
        for row_idx in range(start_row, end_row + 1):
            row_urls = url_data.get(row_idx, {})
            
            for url_col in url_columns:
                url = row_urls.get(url_col)
                if not url:
                    continue
                
                result = extract_data.get(url)
                if not result:
                    continue
                
                matched_count += 1
                
                for field in result_fields:
                    col_name = f"{url_col}-{field}"
                    if col_name in header_map:
                        col_idx = header_map[col_name]
                        value = result.get(field, "")
                        # 清理非法字符后写入
                        ws_write.cell(row=row_idx, column=col_idx, value=clean_illegal_characters(value))
            
            if (row_idx - start_row + 1) % 100 == 0:
                progress = (row_idx - start_row + 1) / total_rows * 100
                log_print(f"处理进度: {row_idx - start_row + 1}/{total_rows} ({progress:.1f}%)")
    
    # 调整新列的列宽
    header_map = read_header_mapping(ws_write, args.header_row)
    for url_col in url_columns:
        for field in result_fields:
            col_name = f"{url_col}-{field}"
            if col_name in header_map:
                col_idx = header_map[col_name]
                col_letter = get_column_letter(col_idx)
                ws_write.column_dimensions[col_letter].width = 20
    
    # 保存
    log_print("正在保存输出文件...")
    wb_write.save(output_path)
    wb_write.close()
    
    # 显示统计
    log_print("=" * 70)
    log_print("任务完成")
    log_print(f"  处理行数: {total_rows}")
    log_print(f"  匹配成功: {matched_count} 个 URL")
    log_print(f"  输出文件: {output_path}")
    log_print("=" * 70)


if __name__ == "__main__":
    main()

