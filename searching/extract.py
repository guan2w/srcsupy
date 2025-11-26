#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract.py

读取 Excel 输入文件，对指定 URL 列调用 ScrapingBee AI Extract 接口提取数据，
将结果记录到 CSV 文件。

依赖：pip install openpyxl scrapingbee

使用示例：
    python extract.py \
        --input-file=/path/to/file.xlsx \
        --sheet-name=Sheet1 \
        --header-row=2 \
        --rows=3+ \
        --url-columns=来源链接1,来源链接2 \
        --concurrency=5
"""

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from scrapingbee import ScrapingBeeClient

# 兼容不同 Python 版本的 toml 库导入
try:
    import tomllib
except ImportError:
    try:
        import tomli as tomllib
    except ImportError:
        tomllib = None


# ==================== 工具函数 ====================

def log_print(*args, level="INFO", **kwargs):
    """通用日志打印函数，带时间戳和级别标识"""
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] [{level}]", *args, **kwargs)


def load_env_file(path: str = ".env"):
    """加载 .env 文件到环境变量"""
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key, value = key.strip(), value.strip()
                if (value.startswith('"') and value.endswith('"')) or \
                   (value.startswith("'") and value.endswith("'")):
                    value = value[1:-1]
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception as e:
        log_print(f"加载 .env 文件出错：{e}", level="WARN")


def load_config(config_path: str) -> Dict[str, Any]:
    """读取配置，默认值兜底"""
    cfg = {
        "timeout_seconds": 120,
        "concurrency": 1,
        "retry_times": 1,
        "proxy": None
    }
    if not os.path.exists(config_path) or tomllib is None:
        return cfg
    try:
        with open(config_path, "rb") as f:
            data = tomllib.load(f)
            bee_cfg = data.get("scrapingbee", {})
            cfg.update({k: v for k, v in bee_cfg.items() if k in cfg})
    except Exception as e:
        log_print(f"解析 config.toml 出错，使用默认配置。错误：{e}", level="WARN")
    return cfg


def parse_rows_spec(spec: str, max_row: int, start_row: int) -> Tuple[int, int]:
    """解析行范围参数"""
    spec = spec.strip()
    if spec.endswith("+"):
        row_start = int(spec[:-1])
        row_end = max_row
    elif "-" in spec:
        parts = spec.split("-", 1)
        row_start, row_end = int(parts[0]), int(parts[1])
    else:
        row_start = int(spec)
        row_end = max_row
    
    if row_start < start_row:
        raise ValueError(f"rows 起始行 {row_start} 不能小于数据起始行 {start_row}")
    if row_end < row_start:
        raise ValueError(f"rows 范围非法：{row_start}-{row_end}")
    
    return row_start, min(row_end, max_row)


def parse_url_columns(spec: str) -> List[str]:
    """解析 URL 列参数"""
    result = []
    for item in spec.split(","):
        item = item.strip()
        if item:
            result.append(item)
    if not result:
        raise ValueError("url-columns 参数不能为空")
    return result


# ==================== Excel 处理 ====================

def read_meta_value(wb, key: str, meta_sheet: str = "meta") -> str:
    """
    从 meta sheet 读取指定 key 的 value
    meta 表结构：A1="key", B1="value"，数据从第2行开始
    """
    if meta_sheet not in wb.sheetnames:
        raise ValueError(f"未找到 meta sheet: {meta_sheet}")
    
    ws = wb[meta_sheet]
    
    # 遍历查找 key
    for row in range(2, ws.max_row + 1):
        cell_key = ws.cell(row=row, column=1).value
        if cell_key and str(cell_key).strip() == key:
            cell_value = ws.cell(row=row, column=2).value
            if not cell_value:
                raise ValueError(f"meta 表中 key='{key}' 的 value 为空，请填写")
            return str(cell_value).strip()
    
    raise ValueError(f"meta 表中未找到 key='{key}'，请添加")


def read_extract_rules_template(wb) -> str:
    """从 meta sheet 读取提取规则模板（key='ai_extract_rules'）"""
    rules_str = read_meta_value(wb, "ai_extract_rules")
    
    # 验证 JSON 格式
    try:
        rules = json.loads(rules_str)
        if not isinstance(rules, dict):
            raise ValueError("提取规则必须是 JSON 对象")
    except json.JSONDecodeError as e:
        raise ValueError(f"提取规则 JSON 解析失败: {e}")
    
    return rules_str


def extract_template_variables(template: str) -> List[str]:
    """从模板中提取变量名，如 {{姓名}} 中的 '姓名'"""
    pattern = r'\{\{(\w+)\}\}'
    return list(set(re.findall(pattern, template)))


def render_rules_template(template: str, row_data: Dict[str, Any]) -> Dict[str, str]:
    """根据行数据渲染提取规则模板"""
    result = template
    for key, value in row_data.items():
        placeholder = "{{" + key + "}}"
        if placeholder in result:
            str_value = str(value).strip() if value is not None else ""
            result = result.replace(placeholder, str_value)
    
    # 解析渲染后的 JSON
    return json.loads(result)


def read_row_data(ws, row_idx: int, header_map: Dict[str, int], variables: List[str]) -> Dict[str, Any]:
    """读取指定行的数据（仅读取模板需要的变量）"""
    row_data = {}
    for var in variables:
        if var in header_map:
            col_idx = header_map[var]
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data[var] = cell_value
        else:
            row_data[var] = ""
    return row_data


def read_header_mapping(ws, header_row: int) -> Dict[str, int]:
    """读取表头，返回列名到列索引的映射"""
    header_map = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value:
            header_map[str(cell_value).strip()] = col
    return header_map


# ==================== ScrapingBee AI Extract ====================

# ScrapingBee timeout 限制：1000-141000 毫秒
SCRAPINGBEE_MAX_TIMEOUT_MS = 141000
SCRAPINGBEE_MAX_TIMEOUT_SEC = 141


def decode_unicode_string(s: str) -> str:
    """解码 Unicode 转义字符串，如 \\u7f51\\u9875 -> 网页"""
    if not isinstance(s, str):
        return s
    try:
        # 检测是否包含 Unicode 转义序列
        if '\\u' in s:
            return s.encode('utf-8').decode('unicode_escape')
        return s
    except:
        return s


def decode_unicode_keys(data: Any) -> Any:
    """递归解码字典中所有 Unicode 转义的 key 和 value"""
    if isinstance(data, dict):
        return {
            decode_unicode_string(k): decode_unicode_keys(v)
            for k, v in data.items()
        }
    elif isinstance(data, list):
        return [decode_unicode_keys(item) for item in data]
    elif isinstance(data, str):
        return decode_unicode_string(data)
    return data


def extract_from_url(
    client: ScrapingBeeClient,
    url: str,
    rules: Dict[str, str],
    timeout: int
) -> Tuple[Optional[Dict], Optional[str], float]:
    """
    调用 ScrapingBee AI Extract 接口
    返回: (提取结果字典, 错误信息, 耗时秒数)
    """
    # 限制 timeout 不超过 ScrapingBee API 限制
    api_timeout_sec = min(timeout, SCRAPINGBEE_MAX_TIMEOUT_SEC)
    api_timeout_ms = api_timeout_sec * 1000
    
    start = time.monotonic()
    try:
        response = client.get(
            url,
            params={
                "ai_extract_rules": rules,
                "timeout": api_timeout_ms
            },
            timeout=api_timeout_sec + 30  # 留出额外时间
        )
        duration = time.monotonic() - start
        
        # 检查响应
        if response.status_code != 200:
            return None, f"HTTP {response.status_code}: {response.text[:500]}", duration
        
        # 解析 JSON 响应
        response_text = response.text if hasattr(response, 'text') else response.content.decode('utf-8')
        
        try:
            # 直接用 json.loads 解析（自动处理 Unicode 转义）
            data = json.loads(response_text)
            # 额外处理可能残留的 Unicode 转义
            data = decode_unicode_keys(data)
            return data, None, duration
        except json.JSONDecodeError as e:
            return None, f"JSON 解析失败: {e}. 响应内容: {response_text[:500]}", duration
            
    except Exception as e:
        return None, f"请求异常: {e}", time.monotonic() - start


def retry_extract(
    client: ScrapingBeeClient,
    url: str,
    rules: Dict[str, str],
    timeout: int,
    retry_times: int
) -> Tuple[Optional[Dict], Optional[str], float]:
    """带重试的提取"""
    last_error = None
    total_duration = 0
    
    for attempt in range(1, retry_times + 2):
        result, error, duration = extract_from_url(client, url, rules, timeout)
        total_duration += duration
        
        if error is None:
            return result, None, total_duration
        
        last_error = error
        if attempt <= retry_times:
            log_print(f"提取失败，尝试 {attempt}/{retry_times + 1}，错误: {error}", level="WARN")
            time.sleep(min(1.0 * attempt, 5.0))
    
    return None, last_error, total_duration


# ==================== 日志处理 ====================

def get_log_header(rules: Dict[str, str]) -> List[str]:
    """生成日志文件表头"""
    base_header = ["url_column", "row_number", "url", "ai_extract_rules", "extract_time", "duration_ms", "SUCCESS", "ERROR"]
    # 添加提取规则中的字段
    rule_fields = list(rules.keys())
    return base_header + rule_fields


def ensure_log_header(log_path: str, header: List[str]):
    """创建日志文件表头"""
    if not os.path.exists(log_path):
        with open(log_path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(header)


def load_existing_urls(log_path: str) -> Set[str]:
    """加载已提取的 URL 集合"""
    existing = set()
    if not os.path.exists(log_path):
        return existing
    
    try:
        with open(log_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = row.get("url", "").strip()
                if url:
                    existing.add(url)
    except Exception as e:
        log_print(f"读取现有日志出错：{e}", level="WARN")
    
    return existing


def write_extract_result(
    writer,
    url: str,
    url_column: str,
    row_number: int,
    rules_used: Dict[str, str],
    extract_time: str,
    duration_ms: int,
    result: Optional[Dict],
    error: Optional[str],
    rule_fields: List[str]
):
    """写入提取结果到日志"""
    # 将使用的规则序列化为 JSON 字符串
    rules_json = json.dumps(rules_used, ensure_ascii=False)
    
    row_data = [
        url_column,
        row_number,
        url,
        rules_json,
        extract_time,
        duration_ms,
        "true" if error is None else "false",
        error or ""
    ]
    
    # 添加提取结果字段
    for field in rule_fields:
        if result and field in result:
            row_data.append(result[field])
        else:
            row_data.append("")
    
    writer.writerow(row_data)


# ==================== 主程序 ====================

def main():
    parser = argparse.ArgumentParser(description="ScrapingBee AI Extract 批量提取脚本")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet-name", required=True, help="输入数据的 Sheet 名称")
    parser.add_argument("--header-row", type=int, default=1, help="表头行号（默认 1）")
    parser.add_argument("--rows", required=True, help="数据行范围，例如 3+ 或 3-100")
    parser.add_argument("--url-columns", required=True, help="URL 列名，多个用逗号分隔，如 来源链接1,来源链接2")
    parser.add_argument("--concurrency", type=int, default=0, help="并发数（0 使用配置文件值）")
    parser.add_argument("--debug", action="store_true", help="输出调试信息")
    
    args = parser.parse_args()
    
    # 验证输入文件
    input_path = os.path.abspath(args.input_file)
    if not os.path.exists(input_path):
        log_print(f"输入文件不存在: {input_path}", level="ERROR")
        sys.exit(1)
    
    # 路径处理
    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    log_path = os.path.join(base_dir, f"{base_name}-extract-log.csv")
    
    # 配置文件查找
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config.toml")
    if not os.path.exists(config_path):
        config_path = os.path.join(base_dir, "config.toml")
    
    # 加载环境变量
    load_env_file(os.path.join(script_dir, ".env"))
    load_env_file(os.path.join(base_dir, ".env"))
    load_env_file()
    
    # 加载配置
    cfg = load_config(config_path)
    
    # 检查并限制 timeout（ScrapingBee API 限制 1-141 秒）
    if cfg["timeout_seconds"] > SCRAPINGBEE_MAX_TIMEOUT_SEC:
        log_print(f"timeout_seconds={cfg['timeout_seconds']} 超过 ScrapingBee 限制，自动调整为 {SCRAPINGBEE_MAX_TIMEOUT_SEC} 秒", level="WARN")
        cfg["timeout_seconds"] = SCRAPINGBEE_MAX_TIMEOUT_SEC
    
    # 并发数优先使用命令行参数
    concurrency = args.concurrency if args.concurrency > 0 else cfg["concurrency"]
    concurrency = max(1, concurrency)
    
    # API 密钥
    api_key = os.environ.get("SCRAPINGBEE_API_KEY", "").strip()
    if not api_key:
        log_print("未找到 SCRAPINGBEE_API_KEY 环境变量", level="ERROR")
        sys.exit(1)
    
    # 解析 URL 列
    try:
        url_columns = parse_url_columns(args.url_columns)
    except ValueError as e:
        log_print(f"url-columns 参数错误: {e}", level="ERROR")
        sys.exit(1)
    
    # 显示配置信息
    log_print("=" * 70)
    log_print("配置信息")
    log_print(f"  输入文件: {input_path}")
    log_print(f"  Sheet: {args.sheet_name}")
    log_print(f"  表头行: {args.header_row}")
    log_print(f"  数据行范围: {args.rows}")
    log_print(f"  URL 列: {url_columns}")
    log_print(f"  并发数: {concurrency}")
    log_print(f"  日志文件: {log_path}")
    log_print(f"  配置文件: {config_path}")
    log_print("=" * 70)
    
    # 加载 Excel 文件
    log_print("正在加载 Excel 文件...")
    wb = load_workbook(input_path, read_only=True, data_only=True)
    
    # 读取提取规则模板
    rules_template = read_extract_rules_template(wb)
    # 解析一次获取字段列表（用于日志表头）
    rules_sample = json.loads(rules_template)
    rule_fields = list(rules_sample.keys())
    log_print(f"提取规则字段: {rule_fields}")
    
    # 提取模板变量
    template_variables = extract_template_variables(rules_template)
    if template_variables:
        log_print(f"模板变量: {template_variables}")
    
    # 验证 sheet
    if args.sheet_name not in wb.sheetnames:
        log_print(f"未找到 sheet: {args.sheet_name}", level="ERROR")
        sys.exit(1)
    
    ws = wb[args.sheet_name]
    
    # 读取表头映射
    header_map = read_header_mapping(ws, args.header_row)
    log_print(f"表头列映射: {header_map}")
    
    # 验证 URL 列是否存在
    missing_cols = [col for col in url_columns if col not in header_map]
    if missing_cols:
        log_print(f"URL 列在表头中未找到: {missing_cols}", level="ERROR")
        sys.exit(1)
    
    # 验证模板变量是否在表头中
    missing_vars = [v for v in template_variables if v not in header_map]
    if missing_vars:
        log_print(f"模板变量在表头中未找到: {missing_vars}", level="ERROR")
        sys.exit(1)
    
    # 解析行范围
    data_start_row = args.header_row + 1
    start_row, end_row = parse_rows_spec(args.rows, ws.max_row, data_start_row)
    log_print(f"处理行范围: {start_row}-{end_row} (共 {end_row - start_row + 1} 行)")
    
    # 初始化日志文件
    log_header = get_log_header(rules_sample)
    ensure_log_header(log_path, log_header)
    
    # 加载已提取的 URL
    existing_urls = load_existing_urls(log_path)
    log_print(f"已存在的 URL 记录: {len(existing_urls)} 条")
    
    # 构建待处理任务：(row_idx, url_column, url, row_data)
    tasks = []
    for row_idx in range(start_row, end_row + 1):
        # 读取该行的模板变量数据（如果有模板变量的话）
        row_data = read_row_data(ws, row_idx, header_map, template_variables) if template_variables else {}
        
        for url_col in url_columns:
            col_idx = header_map[url_col]
            url = ws.cell(row=row_idx, column=col_idx).value
            
            if not url or not str(url).strip():
                if args.debug:
                    log_print(f"行 {row_idx} 列 {url_col} URL 为空，跳过", level="INFO")
                continue
            
            url = str(url).strip()
            
            if url in existing_urls:
                if args.debug:
                    log_print(f"URL 已存在，跳过: {url[:50]}...", level="DEBUG")
                continue
            
            tasks.append((row_idx, url_col, url, row_data))
            existing_urls.add(url)  # 避免同一批次重复处理相同 URL
    
    wb.close()
    
    if not tasks:
        log_print("所有 URL 已处理完成，无需提取")
        return
    
    log_print(f"待处理任务数: {len(tasks)}")
    
    # 创建 ScrapingBee 客户端
    client = ScrapingBeeClient(api_key=api_key)
    
    # 线程安全的日志写入
    log_lock = Lock()
    log_file = open(log_path, "a", newline="", encoding="utf-8")
    log_writer = csv.writer(log_file)
    
    # 进度统计
    processed = 0
    success_count = 0
    error_count = 0
    stats_lock = Lock()
    
    def process_task(task: Tuple[int, str, str, Dict[str, Any]]):
        nonlocal processed, success_count, error_count
        
        row_idx, url_col, url, row_data = task
        
        # 渲染提取规则（支持动态插值）
        if template_variables:
            rules = render_rules_template(rules_template, row_data)
        else:
            rules = rules_sample
        
        # 执行提取
        result, error, duration = retry_extract(
            client, url, rules, cfg["timeout_seconds"], cfg["retry_times"]
        )
        
        extract_time = dt.datetime.now().astimezone().isoformat()
        duration_ms = int(duration * 1000)
        
        # 写入日志（线程安全）
        with log_lock:
            write_extract_result(
                log_writer, url, url_col, row_idx, rules, extract_time, duration_ms,
                result, error, rule_fields
            )
            log_file.flush()
        
        # 更新统计
        with stats_lock:
            processed += 1
            if error:
                error_count += 1
            else:
                success_count += 1
            
            # 进度显示
            percentage = processed / len(tasks) * 100
            status = "✗" if error else "✓"
            log_print(f"[{processed}/{len(tasks)}] ({percentage:.1f}%) {status} 行 {row_idx} [{url_col}]: {url[:40]}...")
    
    # 并发执行
    log_print(f"开始提取，并发数: {concurrency}")
    
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = [executor.submit(process_task, task) for task in tasks]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                log_print(f"任务执行异常: {e}", level="ERROR")
    
    # 清理
    log_file.close()
    
    # 显示统计
    log_print("=" * 70)
    log_print("任务完成")
    log_print(f"  总任务: {len(tasks)}")
    log_print(f"  成功: {success_count}")
    log_print(f"  失败: {error_count}")
    log_print(f"  日志文件: {log_path}")
    log_print("=" * 70)


if __name__ == "__main__":
    main()

